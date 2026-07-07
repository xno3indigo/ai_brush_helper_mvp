from __future__ import annotations

import io
import json
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from ai_brush_helper.common import clean_text, normalize_token_text, read_csv, safe_float, write_csv


CHART_NS = {"c": "http://schemas.openxmlformats.org/drawingml/2006/chart"}
C_NS_URI = "http://schemas.openxmlformats.org/drawingml/2006/chart"

ET.register_namespace("c", C_NS_URI)


def qn(namespace: str, tag: str) -> str:
    return f"{{{namespace}}}{tag}"


def load_mapping_spec(spec_path: Path) -> dict[str, Any]:
    return json.loads(spec_path.read_text(encoding="utf-8"))


def strip_wave(value: Any) -> str:
    text = normalize_token_text(value)
    return re.sub(r"(?:20)?\d{2}[wqh][1-4]|[whq][1-4]$", "", text)


def contains_non_target_wave(value: Any, target_wave: str) -> bool:
    text = normalize_token_text(value)
    wave_tokens = re.findall(r"(?:20)?\d{2}[wqh][1-4]|[whq][1-4]", text)
    if not wave_tokens:
        return False
    target = normalize_token_text(target_wave)
    return all(token not in target and target not in token for token in wave_tokens)


def source_indexes(rows: list[dict[str, str]]) -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    by_label: dict[str, dict[str, str]] = {}
    column_by_norm: dict[str, str] = {}
    if not rows:
        return by_label, column_by_norm
    for column in rows[0].keys():
        column_by_norm[normalize_token_text(column)] = column
    for row in rows:
        label = clean_text(row.get("label", ""), 500)
        if label:
            by_label[strip_wave(label)] = row
    return by_label, column_by_norm


def source_value(
    source_rows: list[dict[str, str]],
    row_key: str,
    col_key: str,
    binding: dict[str, Any],
) -> Any | None:
    if not source_rows:
        return None
    if contains_non_target_wave(row_key, binding.get("target_wave", "")):
        return None
    by_label, column_by_norm = source_indexes(source_rows)
    row_norm = strip_wave(row_key)
    col_norm = strip_wave(col_key)
    segment = clean_text(binding.get("segment_filter", "Total"), 200)
    segment_norm = normalize_token_text(segment)

    # Direct wide table lookup: template row == source label, template col == source column.
    if row_norm in by_label and col_norm in column_by_norm:
        return by_label[row_norm].get(column_by_norm[col_norm])

    # Template row is a source label and template column is the current wave.
    if row_norm in by_label and binding.get("target_wave") and normalize_token_text(binding["target_wave"]) in normalize_token_text(col_key):
        column = column_by_norm.get(segment_norm) or column_by_norm.get("total")
        if column:
            return by_label[row_norm].get(column)

    # Template row is a source label and template column is the metric represented by the source table.
    if row_norm in by_label and metric_key_matches_source(col_key, binding):
        column = column_by_norm.get(segment_norm) or column_by_norm.get("total")
        if column:
            return by_label[row_norm].get(column)

    # Single-series chart: template row == source label, value should come from Total.
    if row_norm in by_label and normalize_token_text(col_key) in {"系列1", "series1", "值", "value"}:
        column = column_by_norm.get(segment_norm) or column_by_norm.get("total")
        if column:
            return by_label[row_norm].get(column)

    # Transposed common case: template col == source label, template row is current-wave metric.
    if col_norm in by_label:
        column = column_by_norm.get(segment_norm) or column_by_norm.get("total")
        if column and metric_row_matches_source(row_key, binding):
            return by_label[col_norm].get(column)

    return None


def metric_key_matches_source(metric_key: str, binding: dict[str, Any]) -> bool:
    key_tokens = normalize_token_text(metric_key)
    title = normalize_token_text(binding.get("source_question_title", ""))
    code = normalize_token_text(binding.get("source_question_code", ""))
    rules = [
        ("b1a", ["第一提及", "tom"]),
        ("b1b", ["其他自发", "自发提及"]),
        ("b1c", ["总知晓", "知晓"]),
        ("c1a", ["处方过", "p3m处方过"]),
        ("c1b", ["经常处方"]),
        ("c1c", ["最常处方"]),
        ("c2", ["份额", "share", "soc"]),
        ("nps", ["nps", "promotor", "detractor"]),
        ("e1a", ["覆盖", "sov"]),
        ("e2", ["频率", "次数"]),
        ("f1", ["活动", "会议", "接触次数"]),
    ]
    for source_code, aliases in rules:
        source_hit = source_code in code or source_code in title
        metric_hit = any(normalize_token_text(alias) in key_tokens for alias in aliases)
        if source_hit and metric_hit:
            return True
    return False


def metric_row_matches_source(row_key: str, binding: dict[str, Any]) -> bool:
    if contains_non_target_wave(row_key, binding.get("target_wave", "")):
        return False
    row_tokens = normalize_token_text(row_key)
    title = normalize_token_text(binding.get("source_question_title", ""))
    code = normalize_token_text(binding.get("source_question_code", ""))
    if code and code in row_tokens:
        return True
    metric_aliases = {
        "第一提及": ["第一提及", "tom"],
        "提示后知晓": ["总知晓", "知晓"],
        "处方使用": ["处方过", "p3m"],
        "经常处方": ["经常处方"],
        "最常处方": ["最常处方"],
        "nps": ["nps"],
        "拜访覆盖": ["覆盖"],
        "拜访次数": ["频率", "次数"],
    }
    for source_keyword, row_keywords in metric_aliases.items():
        if normalize_token_text(source_keyword) in title and any(normalize_token_text(keyword) in row_tokens for keyword in row_keywords):
            return True
    # If the template row only says current wave, allow one-metric source tables.
    return bool(binding.get("target_wave") and normalize_token_text(binding["target_wave"]) in row_tokens)


def project_source_to_workbook(
    workbook_bytes: bytes,
    source_rows: list[dict[str, str]],
    binding: dict[str, Any],
) -> tuple[bytes, dict[str, Any], list[list[Any]]]:
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    sheet = workbook[workbook.sheetnames[0]]
    changed = 0
    considered = 0
    sample_changes: list[str] = []

    for row_index in range(2, sheet.max_row + 1):
        row_key = clean_text(sheet.cell(row_index, 1).value, 300)
        if not row_key:
            continue
        for col_index in range(2, sheet.max_column + 1):
            col_key = clean_text(sheet.cell(1, col_index).value, 300)
            if not col_key:
                continue
            considered += 1
            value = source_value(source_rows, row_key, col_key, binding)
            number = safe_float(value)
            if number is None:
                continue
            sheet.cell(row_index, col_index, value=number)
            changed += 1
            if len(sample_changes) < 10:
                sample_changes.append(f"R{row_index}C{col_index}={number:g}")

    matrix: list[list[Any]] = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True):
        matrix.append(list(row))

    output = io.BytesIO()
    workbook.save(output)
    return (
        output.getvalue(),
        {
            "workbook_cells_considered": considered,
            "workbook_cells_updated": changed,
            "sample_changes": sample_changes,
        },
        matrix,
    )


def set_cache_points(cache: ET.Element, values: list[Any], numeric: bool) -> None:
    for child in list(cache):
        if child.tag in {qn(C_NS_URI, "ptCount"), qn(C_NS_URI, "pt")}:
            cache.remove(child)
    cache.insert(0, ET.Element(qn(C_NS_URI, "ptCount"), {"val": str(len(values))}))
    for index, value in enumerate(values):
        pt = ET.Element(qn(C_NS_URI, "pt"), {"idx": str(index)})
        v = ET.SubElement(pt, qn(C_NS_URI, "v"))
        if numeric:
            number = safe_float(value)
            v.text = "" if number is None else f"{number:g}"
        else:
            v.text = "" if value is None else str(value)
        cache.insert(index + 1, pt)


def first_cache(container: ET.Element | None) -> ET.Element | None:
    if container is None:
        return None
    cache = container.find(".//c:strCache", CHART_NS)
    if cache is not None:
        return cache
    return container.find(".//c:numCache", CHART_NS)


def update_chart_xml_from_matrix(chart_bytes: bytes, matrix: list[list[Any]]) -> tuple[bytes, dict[str, Any]]:
    root = ET.fromstring(chart_bytes)
    series = root.findall(".//c:ser", CHART_NS)
    if not series:
        return chart_bytes, {"chart_cache_status": "unsupported_no_series", "chart_series_updated": 0}
    if not matrix or len(matrix) < 2:
        return chart_bytes, {"chart_cache_status": "missing_matrix", "chart_series_updated": 0}

    headers = [clean_text(value, 200) for value in matrix[0][1:]]
    categories = [row[0] for row in matrix[1:] if row and clean_text(row[0], 200)]
    updated = 0
    for index, ser in enumerate(series):
        if index >= len(headers):
            break
        values = [row[index + 1] if index + 1 < len(row) else "" for row in matrix[1 : 1 + len(categories)]]
        tx_cache = first_cache(ser.find("c:tx", CHART_NS))
        if tx_cache is not None:
            set_cache_points(tx_cache, [headers[index]], numeric=False)
        cat_cache = first_cache(ser.find("c:cat", CHART_NS))
        if cat_cache is not None:
            set_cache_points(cat_cache, categories, numeric=False)
        val_cache = first_cache(ser.find("c:val", CHART_NS))
        if val_cache is None:
            continue
        set_cache_points(val_cache, values, numeric=True)
        updated += 1

    status = "ok" if updated else "unsupported_no_value_cache"
    if 0 < updated < len(series):
        status = "partial_series_count_mismatch"
    return (
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
        {"chart_cache_status": status, "chart_series_updated": updated, "chart_series": len(series)},
    )


def render_enhanced(
    pptx_path: Path,
    import_dir: Path,
    mapping_spec: Path,
    output_pptx: Path,
    out_dir: Path,
    min_confidence: str = "medium",
) -> dict[str, Any]:
    confidence_rank = {"no_match": 0, "low": 1, "medium": 2, "high": 3}
    minimum = confidence_rank.get(min_confidence, 2)
    spec = load_mapping_spec(mapping_spec)
    replacements: dict[str, bytes] = {}
    validation_rows: list[dict[str, Any]] = []

    with zipfile.ZipFile(pptx_path, "r") as archive:
        for binding in spec.get("bindings", []):
            status = "skipped"
            detail: dict[str, Any] = {}
            source_file = clean_text(binding.get("source_file", ""), 500)
            workbook_path = clean_text(binding.get("embedded_workbook", ""), 500)
            chart_path = clean_text(binding.get("chart_path", ""), 500)
            confidence = clean_text(binding.get("confidence", ""), 50)

            if confidence_rank.get(confidence, 0) < minimum:
                status = "skipped_low_confidence"
            elif not source_file:
                status = "missing_source_table"
            elif not workbook_path or workbook_path not in archive.namelist():
                status = "missing_embedded_workbook"
            elif workbook_path.lower().endswith(".xlsb"):
                status = "unsupported_xlsb"
            else:
                source_path = import_dir / source_file
                if not source_path.exists():
                    status = "missing_source_file"
                else:
                    source_rows = read_csv(source_path)
                    workbook_bytes, workbook_report, matrix = project_source_to_workbook(
                        archive.read(workbook_path),
                        source_rows,
                        binding,
                    )
                    detail.update(workbook_report)
                    if workbook_report["workbook_cells_updated"] <= 0:
                        status = "no_matching_cells"
                    else:
                        replacements[workbook_path] = workbook_bytes
                        if chart_path in archive.namelist():
                            chart_bytes, chart_report = update_chart_xml_from_matrix(archive.read(chart_path), matrix)
                            replacements[chart_path] = chart_bytes
                            detail.update(chart_report)
                            status = chart_report["chart_cache_status"]
                        else:
                            status = "workbook_updated_chart_missing"

            validation_rows.append(
                {
                    "page": binding.get("page", ""),
                    "chart_sort": binding.get("chart_sort", ""),
                    "chart_path": chart_path,
                    "embedded_workbook": workbook_path,
                    "source_table": binding.get("source_table", ""),
                    "confidence": confidence,
                    "status": status,
                    "detail": json.dumps(detail, ensure_ascii=False),
                }
            )

        output_pptx.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(output_pptx, "w", zipfile.ZIP_DEFLATED) as output:
            for item in archive.infolist():
                output.writestr(item, replacements.get(item.filename, archive.read(item.filename)))

    write_csv(
        out_dir / "render_validation.enhanced.csv",
        validation_rows,
        ["page", "chart_sort", "chart_path", "embedded_workbook", "source_table", "confidence", "status", "detail"],
    )
    status_counts: dict[str, int] = {}
    for row in validation_rows:
        status_counts[row["status"]] = status_counts.get(row["status"], 0) + 1
    report = {
        "input_pptx": str(pptx_path),
        "output_pptx": str(output_pptx),
        "mapping_spec": str(mapping_spec),
        "bindings": len(spec.get("bindings", [])),
        "updated_files": len(replacements),
        "status_counts": status_counts,
        "min_confidence": min_confidence,
    }
    (out_dir / "render_report.enhanced.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report
