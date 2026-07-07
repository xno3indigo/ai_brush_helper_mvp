#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import json
import posixpath
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import pymysql
from openpyxl import load_workbook, Workbook


PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
C_NS_URI = "http://schemas.openxmlformats.org/drawingml/2006/chart"
A_NS_URI = "http://schemas.openxmlformats.org/drawingml/2006/main"
R_NS_URI = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
CHART_NS = {"c": C_NS_URI}
DRAWING_NS = {"a": A_NS_URI}

ET.register_namespace("c", C_NS_URI)
ET.register_namespace("a", A_NS_URI)
ET.register_namespace("r", R_NS_URI)

BRAND_ORDER = ["百泽安", "达伯舒", "艾瑞卡", "拓益", "可瑞达", "欧狄沃"]
FULL_BRAND_ORDER = ["百泽安", "达伯舒", "艾瑞卡", "拓益", "可瑞达", "欧狄沃", "汉斯状", "择捷美"]


def qn(namespace: str, tag: str) -> str:
    return f"{{{namespace}}}{tag}"


def clean(value: Any) -> str:
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def norm(value: Any) -> str:
    return re.sub(r"[\s（）()_\-+/]+", "", clean(value).lower())


def connect(args: argparse.Namespace) -> pymysql.connections.Connection:
    return pymysql.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        database=args.database,
        charset="utf8mb4",
    )


def fetch_table(conn: pymysql.connections.Connection, table: str) -> list[dict[str, Any]]:
    with conn.cursor(pymysql.cursors.DictCursor) as cur:
        cur.execute(f"select * from `{table}`")
        return list(cur.fetchall())


def brand_index(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        brand = row.get("brand", "")
        for short in FULL_BRAND_ORDER:
            if norm(short) in norm(brand):
                result[short] = row
    return result


def relationship_targets(archive: zipfile.ZipFile, rels_path: str, base_dir: str) -> dict[str, str]:
    if rels_path not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read(rels_path))
    result: dict[str, str] = {}
    for rel in root.findall(f".//{{{PACKAGE_REL_NS}}}Relationship"):
        rel_id = rel.attrib.get("Id", "")
        target = rel.attrib.get("Target", "")
        if rel_id and target:
            result[rel_id] = posixpath.normpath(posixpath.join(base_dir, target))
    return result


def slide_chart_paths(archive: zipfile.ZipFile, slide_no: int) -> list[str]:
    slide_path = f"ppt/slides/slide{slide_no}.xml"
    root = ET.fromstring(archive.read(slide_path))
    rels = relationship_targets(archive, f"ppt/slides/_rels/slide{slide_no}.xml.rels", "ppt/slides")
    paths = []
    for chart in root.findall(".//c:chart", {"c": C_NS_URI, "r": R_NS_URI}):
        rel_id = chart.attrib.get(qn(R_NS_URI, "id"), "")
        if rel_id in rels:
            paths.append(rels[rel_id])
    return paths


def embedded_workbooks(archive: zipfile.ZipFile, chart_path: str) -> list[str]:
    chart_name = posixpath.basename(chart_path)
    rels_path = posixpath.join(posixpath.dirname(chart_path), "_rels", chart_name + ".rels")
    rels = relationship_targets(archive, rels_path, posixpath.dirname(chart_path))
    return [path for path in rels.values() if path.startswith("ppt/embeddings/")]


def workbook_bytes_from_matrix(matrix: list[list[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row_index, row in enumerate(matrix, start=1):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index, value=value)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def safe_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[:\\/?*\[\]]", "_", value)[:31] or "Sheet"
    name = base
    index = 2
    while name in used:
        suffix = f"_{index}"
        name = base[: 31 - len(suffix)] + suffix
        index += 1
    used.add(name)
    return name


def append_matrix_sheet(wb: Workbook, name: str, matrix: list[list[Any]], used: set[str]) -> None:
    ws = wb.create_sheet(safe_sheet_name(name, used))
    for row in matrix:
        ws.append(row)


def rows_to_matrix(rows: list[dict[str, Any]]) -> list[list[Any]]:
    if not rows:
        return [["no_rows"]]
    headers = list(rows[0].keys())
    return [headers] + [[row.get(header) for header in headers] for row in rows]


def write_data_workbook(
    path: Path,
    chart_matrices: list[tuple[str, list[list[Any]]]],
    source_tables: dict[str, list[dict[str, Any]]],
    validation: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    used = {wb.active.title}
    wb.active.title = safe_sheet_name("readme", set())
    used = {wb.active.title}
    wb.active.append(["type", "name", "description"])
    wb.active.append(["chart_matrix", "chart_*", "actual matrices written to embedded PPT workbooks"])
    wb.active.append(["source_table", "source_*", "raw rows fetched from tracking_dlbcl"])
    wb.active.append(["validation", "render_validation", "rendered/skipped status for each target chart"])

    for name, matrix in chart_matrices:
        append_matrix_sheet(wb, name, matrix, used)

    for table, rows in source_tables.items():
        append_matrix_sheet(wb, f"source_{table}", rows_to_matrix(rows), used)

    append_matrix_sheet(wb, "render_validation", rows_to_matrix(validation), used)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def replace_embedded_workbook(original_bytes: bytes, matrix: list[list[Any]]) -> bytes:
    wb = load_workbook(io.BytesIO(original_bytes))
    ws = wb[wb.sheetnames[0]]
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    for row_index, row in enumerate(matrix, start=1):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index, value=value)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def set_cache(cache: ET.Element, values: list[Any], numeric: bool) -> None:
    for child in list(cache):
        if child.tag in {qn(C_NS_URI, "ptCount"), qn(C_NS_URI, "pt")}:
            cache.remove(child)
    cache.insert(0, ET.Element(qn(C_NS_URI, "ptCount"), {"val": str(len(values))}))
    for index, value in enumerate(values):
        pt = ET.Element(qn(C_NS_URI, "pt"), {"idx": str(index)})
        v = ET.SubElement(pt, qn(C_NS_URI, "v"))
        if numeric and value not in (None, ""):
            v.text = f"{float(value):g}"
        else:
            v.text = "" if value is None else str(value)
        cache.insert(index + 1, pt)


def first_cache(container: ET.Element | None) -> ET.Element | None:
    if container is None:
        return None
    cache = container.find(".//c:strCache", CHART_NS)
    if cache is not None:
        return cache
    return container.find(".//c:numCache", CHART_NS)


def update_tx(ser: ET.Element, name: str) -> None:
    tx = ser.find("c:tx", CHART_NS)
    if tx is None:
        tx = ET.SubElement(ser, qn(C_NS_URI, "tx"))
    cache = first_cache(tx)
    if cache is not None:
        set_cache(cache, [name], numeric=False)
        return
    v = tx.find("c:v", CHART_NS)
    if v is None:
        v = ET.SubElement(tx, qn(C_NS_URI, "v"))
    v.text = name


def update_bar_chart(chart_bytes: bytes, categories: list[str], series: list[tuple[str, list[float]]]) -> bytes:
    root = ET.fromstring(chart_bytes)
    ser_nodes = root.findall(".//c:ser", CHART_NS)
    for index, (name, values) in enumerate(series):
        if index >= len(ser_nodes):
            break
        ser = ser_nodes[index]
        update_tx(ser, name)
        cat_cache = first_cache(ser.find("c:cat", CHART_NS))
        if cat_cache is not None:
            set_cache(cat_cache, categories, numeric=False)
        val_cache = first_cache(ser.find("c:val", CHART_NS))
        if val_cache is not None:
            set_cache(val_cache, values, numeric=True)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def update_scatter_chart(chart_bytes: bytes, series: list[tuple[str, list[float], list[float]]]) -> bytes:
    root = ET.fromstring(chart_bytes)
    ser_nodes = root.findall(".//c:ser", CHART_NS)
    parent_by_ser: dict[int, ET.Element] = {}
    for parent in root.iter():
        for child in list(parent):
            if child.tag == qn(C_NS_URI, "ser"):
                parent_by_ser[id(child)] = parent
    for ser in ser_nodes[len(series) :]:
        parent = parent_by_ser.get(id(ser))
        if parent is not None:
            parent.remove(ser)
    ser_nodes = root.findall(".//c:ser", CHART_NS)
    for index, (name, x_values, y_values) in enumerate(series):
        if index >= len(ser_nodes):
            break
        ser = ser_nodes[index]
        update_tx(ser, name)
        x_cache = first_cache(ser.find("c:xVal", CHART_NS))
        y_cache = first_cache(ser.find("c:yVal", CHART_NS))
        if x_cache is not None:
            set_cache(x_cache, x_values, numeric=True)
        if y_cache is not None:
            set_cache(y_cache, y_values, numeric=True)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def update_slide_text(slide_bytes: bytes, replacements: dict[str, str]) -> bytes:
    root = ET.fromstring(slide_bytes)
    for node in root.findall(".//a:t", DRAWING_NS):
        text = node.text or ""
        for src, dst in replacements.items():
            text = text.replace(src, dst)
        node.text = text
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def render(args: argparse.Namespace) -> dict[str, Any]:
    conn = connect(args)
    awareness_rows = fetch_table(conn, "brand_awareness_total_26w2")
    sov_rows = fetch_table(conn, "sov_total_26w2")
    nps_rows = fetch_table(conn, "nps_score_total_26w2")
    equity_source_rows = fetch_table(conn, "brand_equity_total_26w2")
    awareness = brand_index(awareness_rows)
    sov = brand_index(sov_rows)
    nps = brand_index(nps_rows)
    equity_rows = [
        row
        for row in equity_source_rows
        if row.get("ranking") is not None and row.get("attr") and all(row.get(col) is not None for col in [
            "赞必佳（芦比替定）_132",
            "免疫 + 化疗（不含芦比替定_170",
            "福可维（安罗替尼）_161",
            "安泰适（塔拉妥单抗）_112",
        ])
    ]
    equity_rows = [row for row in equity_rows if float(row["ranking"]) <= 10]
    conn.close()

    validation: list[dict[str, Any]] = []
    replacements: dict[str, bytes] = {}
    chart_matrices: list[tuple[str, list[list[Any]]]] = []

    categories = BRAND_ORDER
    tom_values = [float(awareness[brand]["tom"]) if brand in awareness else None for brand in categories]
    sov_values = [float(sov[brand]["sov"]) if brand in sov else None for brand in categories]
    nps_values = [float(nps[brand]["NPS"]) if brand in nps else None for brand in categories]
    # The small horizontal charts on slide 7 are overlaid on static brand labels.
    # Their category axis renders last data point at the top, so chart data must be
    # written bottom-to-top to align visually with the template's top-to-bottom labels.
    row_overlay_categories = list(reversed(categories))
    sov_plot_values = list(reversed(sov_values))
    nps_plot_values = list(reversed(nps_values))

    equity_products = [
        "赞必佳（芦比替定）_132",
        "免疫 + 化疗（不含芦比替定_170",
        "福可维（安罗替尼）_161",
        "安泰适（塔拉妥单抗）_112",
    ]
    equity_names = ["赞必佳", "免疫+化疗", "福可维", "安泰适"]
    x_values = [float(row["ranking"]) for row in equity_rows]
    attr_values = [row["attr"] for row in equity_rows]
    equity_matrix_raw = [["id", *equity_names, "attribute", "ranking"]]
    equity_matrix_pct = [["id", *equity_names, "attribute", "ranking"]]
    for idx, row in enumerate(equity_rows, start=1):
        equity_matrix_raw.append([idx, *[float(row[col]) for col in equity_products], row["attr"], float(row["ranking"])])
        equity_matrix_pct.append([idx, *[float(row[col]) / 100 for col in equity_products], row["attr"], float(row["ranking"])])

    with zipfile.ZipFile(args.pptx, "r") as archive:
        slide7_charts = slide_chart_paths(archive, 7)
        slide8_charts = slide_chart_paths(archive, 8)
        jobs = [
            {
                "slide": 7,
                "sort": 1,
                "chart": slide7_charts[0],
                "matrix": [["", "25W2", "26W2"], *[[brand, None, value] for brand, value in zip(categories, tom_values)]],
                "chart_bytes": update_bar_chart(archive.read(slide7_charts[0]), categories, [("25W2", [None] * len(categories)), ("26W2", tom_values)]),
                "source": "brand_awareness_total_26w2.tom",
                "sheet": "chart_7_1_TOM",
            },
            {
                "slide": 7,
                "sort": 2,
                "chart": slide7_charts[1],
                "matrix": [["status", "reason"], ["skipped", "SOC source table/field not found in tracking_dlbcl"]],
                "chart_bytes": None,
                "source": "SOC source table missing",
                "sheet": "chart_7_2_SOC_skipped",
                "skip_status": "skipped_source_table_not_found",
            },
            {
                "slide": 7,
                "sort": 3,
                "chart": slide7_charts[2],
                "matrix": [["plot_order_brand", "SOV"], *[[brand, value] for brand, value in zip(row_overlay_categories, sov_plot_values)]],
                "chart_bytes": update_bar_chart(archive.read(slide7_charts[2]), row_overlay_categories, [("SOV", sov_plot_values)]),
                "source": "sov_total_26w2.sov",
                "sheet": "chart_7_3_SOV",
            },
            {
                "slide": 7,
                "sort": 4,
                "chart": slide7_charts[3],
                "matrix": [["plot_order_brand", "NPS"], *[[brand, value] for brand, value in zip(row_overlay_categories, nps_plot_values)]],
                "chart_bytes": update_bar_chart(archive.read(slide7_charts[3]), row_overlay_categories, [("NPS", nps_plot_values)]),
                "source": "nps_score_total_26w2.NPS",
                "sheet": "chart_7_4_NPS",
            },
            {
                "slide": 7,
                "sort": 5,
                "chart": slide7_charts[4],
                "matrix": [["status", "reason"], ["skipped", "Adoption Rate source table not found and embedded workbook is xlsb"]],
                "chart_bytes": None,
                "source": "Adoption Rate source table missing",
                "sheet": "chart_7_5_Adoption_skipped",
                "skip_status": "skipped_source_table_not_found_xlsb",
            },
            {
                "slide": 8,
                "sort": 1,
                "chart": slide8_charts[0],
                "matrix": equity_matrix_raw,
                "chart_bytes": update_scatter_chart(
                    archive.read(slide8_charts[0]),
                    [
                        (name, [float(row[col]) for row in equity_rows], x_values)
                        for name, col in zip(equity_names, equity_products)
                    ],
                ),
                "source": "brand_equity_total_26w2 raw score",
                "sheet": "chart_8_1_equity_raw",
            },
            {
                "slide": 8,
                "sort": 2,
                "chart": slide8_charts[1],
                "matrix": equity_matrix_pct,
                "chart_bytes": update_scatter_chart(
                    archive.read(slide8_charts[1]),
                    [
                        (name, [float(row[col]) / 100 for row in equity_rows], x_values)
                        for name, col in zip(equity_names, equity_products)
                    ],
                ),
                "source": "brand_equity_total_26w2 score/100",
                "sheet": "chart_8_2_equity_ratio",
            },
        ]

        for job in jobs:
            chart_path = job["chart"]
            chart_matrices.append((job["sheet"], job["matrix"]))
            if job.get("skip_status"):
                validation.append(
                    {
                        "slide": job["slide"],
                        "chart_sort": job["sort"],
                        "chart_path": chart_path,
                        "source": job["source"],
                        "status": job["skip_status"],
                        "rows": len(job["matrix"]) - 1,
                        "columns": len(job["matrix"][0]),
                    }
                )
                continue
            embeddings = embedded_workbooks(archive, chart_path)
            if not embeddings:
                status = "missing_embedded_workbook"
            else:
                workbook_path = embeddings[0]
                if workbook_path.lower().endswith(".xlsb"):
                    status = "unsupported_xlsb"
                else:
                    replacements[workbook_path] = replace_embedded_workbook(archive.read(workbook_path), job["matrix"])
                    replacements[chart_path] = job["chart_bytes"]
                    status = "ok"
            validation.append(
                {
                    "slide": job["slide"],
                    "chart_sort": job["sort"],
                    "chart_path": chart_path,
                    "source": job["source"],
                    "status": status,
                    "rows": len(job["matrix"]) - 1,
                    "columns": len(job["matrix"][0]),
                }
            )

        for slide_no in [7, 8]:
            slide_path = f"ppt/slides/slide{slide_no}.xml"
            replacements[slide_path] = update_slide_text(
                replacements.get(slide_path, archive.read(slide_path)),
                {"N=190": "N=114", "N=XX": "N=114", "2026W1": "26W2", "26W1": "26W2", "2025W2": "25W2"},
            )

        args.output.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(args.output, "w", zipfile.ZIP_DEFLATED) as output:
            for item in archive.infolist():
                output.writestr(item, replacements.get(item.filename, archive.read(item.filename)))

    args.report.parent.mkdir(parents=True, exist_ok=True)
    with args.report.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["slide", "chart_sort", "chart_path", "source", "status", "rows", "columns"])
        writer.writeheader()
        writer.writerows(validation)
    write_data_workbook(
        args.data_xlsx,
        chart_matrices,
        {
            "brand_awareness_total_26w2": awareness_rows,
            "sov_total_26w2": sov_rows,
            "nps_score_total_26w2": nps_rows,
            "brand_equity_total_26w2": equity_source_rows,
        },
        validation,
    )
    return {
        "output": str(args.output),
        "report": str(args.report),
        "data_xlsx": str(args.data_xlsx),
        "updated_jobs": sum(1 for row in validation if row["status"] == "ok"),
        "validation": validation,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render PPT slides 7/8 from tracking_dlbcl database for local testing.")
    parser.add_argument("--pptx", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--data-xlsx", type=Path, required=True)
    parser.add_argument("--host", default="192.168.20.7")
    parser.add_argument("--port", type=int, default=3306)
    parser.add_argument("--user", default="root")
    parser.add_argument("--password", default="root")
    parser.add_argument("--database", default="tracking_dlbcl")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    print(json.dumps(render(args), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
