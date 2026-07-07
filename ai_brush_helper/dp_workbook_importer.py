from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ai_brush_helper.common import clean_text, ensure_dir, slug, unique_name, write_csv


@dataclass
class ParsedTable:
    table_name: str
    question_code: str
    question_title: str
    source_sheet: str
    start_row: int
    end_row: int
    headers: list[str]
    header_levels: list[list[str]]
    rows: list[list[Any]]


def _cell_matrix(ws: Any) -> list[list[Any]]:
    values = [[cell.value for cell in row] for row in ws.iter_rows()]
    # Fill merged cells so multi-level headers keep their parent labels.
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        value = ws.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                values[row - 1][col - 1] = value
    return values


def _find_table_blocks(values: list[list[Any]]) -> list[tuple[int, int, int]]:
    blocks: list[tuple[int, int, int]] = []
    for row_index, row in enumerate(values):
        for col_index, value in enumerate(row):
            if value == "Table Start":
                end_index = len(values) - 1
                for candidate in range(row_index + 1, len(values)):
                    if any(cell == "Table End" for cell in values[candidate]):
                        end_index = candidate
                        break
                blocks.append((row_index, col_index, end_index))
    return blocks


def _first_non_empty(candidates: list[Any]) -> str:
    for value in candidates:
        text = clean_text(value, 300)
        if text and text != "˙":
            return text
    return ""


def _question_code(title: str, fallback_index: int) -> str:
    compact = clean_text(title, 300).replace(" ", "")
    match = re.match(r"([A-Za-z]+\d+[A-Za-z0-9]*)", compact)
    return match.group(1) if match else f"table{fallback_index:02d}"


def _row_values(row: list[Any], start_col: int, end_col: int) -> list[str]:
    return [clean_text(row[col] if col < len(row) else "", 500) for col in range(start_col, end_col + 1)]


def _parse_block(
    values: list[list[Any]],
    source_sheet: str,
    block_index: int,
    start_row: int,
    marker_col: int,
    end_row: int,
    wave: str,
    used_names: dict[str, int],
) -> ParsedTable | None:
    block = values[start_row + 1 : end_row]
    if not block:
        return None

    title = ""
    for row in block[:3]:
        title = _first_non_empty(row[marker_col : marker_col + 10])
        if title:
            break
    if not title:
        title = _first_non_empty([cell for row in block[:3] for cell in row])

    header_index: int | None = None
    for index, row in enumerate(block):
        row_text = [clean_text(cell, 200) for cell in row]
        if "Sample Size" in row_text and index > 0:
            header_index = index - 1
            break
    if header_index is None:
        for index, row in enumerate(block):
            joined = " ".join(clean_text(cell, 100) for cell in row)
            if "Total" in joined or "Segment" in joined:
                header_index = index
                break
    if header_index is None:
        header_index = 1 if len(block) > 1 else 0

    start_col = max(0, marker_col)
    end_col = start_col
    for row in block[header_index:]:
        for col, value in enumerate(row):
            if col >= start_col and clean_text(value, 200):
                end_col = max(end_col, col)

    header_1 = _row_values(block[header_index], start_col, end_col)
    header_2 = _row_values(block[header_index + 1], start_col, end_col) if header_index + 1 < len(block) else []
    has_second_header = bool(header_2) and header_2[0] != "Sample Size" and any(header_2[1:])

    filled_header_1: list[str] = []
    current = ""
    for value in header_1:
        if value and value != "Segment":
            current = value
        filled_header_1.append(current)

    headers: list[str] = []
    header_levels: list[list[str]] = []
    for offset in range(end_col - start_col + 1):
        h1 = header_1[offset] if offset < len(header_1) else ""
        h1_filled = filled_header_1[offset] if offset < len(filled_header_1) else ""
        h2 = header_2[offset] if offset < len(header_2) else ""
        if offset == 0:
            header = "label"
            levels = ["label", ""]
        elif has_second_header:
            parent = h1_filled or h1
            child = h2
            header = f"{parent}__{child}" if parent and child and parent != child else parent or child or f"col_{offset + 1}"
            levels = [parent, child]
        else:
            header = h1 or f"col_{offset + 1}"
            levels = [header, ""]

        base = header
        suffix = 2
        while header in headers:
            header = f"{base}_{suffix}"
            suffix += 1
        headers.append(header)
        header_levels.append(levels)

    data_start = header_index + (2 if has_second_header else 1)
    data_rows: list[list[Any]] = []
    for row in block[data_start:]:
        row_values = _row_values(row, start_col, end_col)
        if not any(row_values):
            continue
        if row_values[0] in {"Table Start", "Table End"}:
            continue
        data_rows.append(row_values)

    if not data_rows:
        return None

    code = _question_code(title, block_index)
    base_name = f"{slug(code, 20)}_{slug(title, 80)}"
    if wave:
        base_name += f"_{slug(wave, 20)}"
    table_name = unique_name(re.sub(r"_+", "_", base_name).strip("_"), used_names)
    return ParsedTable(
        table_name=table_name,
        question_code=code,
        question_title=title,
        source_sheet=source_sheet,
        start_row=start_row + 1,
        end_row=end_row + 1,
        headers=headers,
        header_levels=header_levels,
        rows=data_rows,
    )


def parse_dp_workbook(xlsx_path: Path, sheet_name: str = "DP_问卷", wave: str = "") -> list[ParsedTable]:
    workbook = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    ws = workbook[sheet_name]
    values = _cell_matrix(ws)
    used_names: dict[str, int] = {}
    tables: list[ParsedTable] = []
    for block_index, (start_row, marker_col, end_row) in enumerate(_find_table_blocks(values), start=1):
        parsed = _parse_block(values, sheet_name, block_index, start_row, marker_col, end_row, wave, used_names)
        if parsed:
            tables.append(parsed)
    return tables


def write_import_outputs(tables: list[ParsedTable], out_dir: Path, wave: str = "") -> dict[str, Any]:
    ensure_dir(out_dir)
    table_dir = out_dir / "result_tables"
    ensure_dir(table_dir)

    manifest: list[dict[str, Any]] = []
    long_rows: list[dict[str, Any]] = []
    for table in tables:
        wide_path = table_dir / f"{table.table_name}.csv"
        with wide_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(table.headers)
            writer.writerows(table.rows)

        manifest.append(
            {
                "table_name": table.table_name,
                "file": str(wide_path.relative_to(out_dir)),
                "question_code": table.question_code,
                "question_title": table.question_title,
                "source_sheet": table.source_sheet,
                "start_row": table.start_row,
                "end_row": table.end_row,
                "row_count": len(table.rows),
                "column_count": len(table.headers),
                "columns": table.headers,
                "header_levels": table.header_levels,
                "wave": wave,
            }
        )

        for source_row_offset, row in enumerate(table.rows, start=1):
            row_label = clean_text(row[0] if row else "", 500)
            for col_index, header in enumerate(table.headers[1:], start=1):
                level_1, level_2 = table.header_levels[col_index] if col_index < len(table.header_levels) else [header, ""]
                long_rows.append(
                    {
                        "table_name": table.table_name,
                        "question_code": table.question_code,
                        "question_title": table.question_title,
                        "row_label": row_label,
                        "column_key": header,
                        "column_level_1": level_1,
                        "column_level_2": level_2,
                        "segment": level_1,
                        "metric": level_2 or row_label,
                        "value": row[col_index] if col_index < len(row) else "",
                        "source_sheet": table.source_sheet,
                        "source_row": table.start_row + source_row_offset,
                        "source_column_index": col_index + 1,
                        "wave": wave,
                    }
                )

    (out_dir / "_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(
        out_dir / "result_tables_long.csv",
        long_rows,
        [
            "table_name",
            "question_code",
            "question_title",
            "row_label",
            "column_key",
            "column_level_1",
            "column_level_2",
            "segment",
            "metric",
            "value",
            "source_sheet",
            "source_row",
            "source_column_index",
            "wave",
        ],
    )
    return {"tables": len(tables), "long_rows": len(long_rows), "out_dir": str(out_dir)}


def import_dp_workbook(xlsx_path: Path, out_dir: Path, sheet_name: str = "DP_问卷", wave: str = "") -> dict[str, Any]:
    tables = parse_dp_workbook(xlsx_path, sheet_name=sheet_name, wave=wave)
    return write_import_outputs(tables, out_dir, wave=wave)

