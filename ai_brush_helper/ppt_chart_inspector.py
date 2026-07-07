from __future__ import annotations

import io
import json
import posixpath
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from ai_brush_helper.common import clean_text, tokenize, write_csv


DRAWING_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
CHART_NS = {"c": "http://schemas.openxmlformats.org/drawingml/2006/chart"}
REL_NS = {"r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
C_NS_URI = "http://schemas.openxmlformats.org/drawingml/2006/chart"
R_NS_URI = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
SLIDE_RE = re.compile(r"ppt/slides/slide(\d+)\.xml$")


def qn(namespace: str, tag: str) -> str:
    return f"{{{namespace}}}{tag}"


def ppt_slide_files(archive: zipfile.ZipFile) -> list[str]:
    files = [name for name in archive.namelist() if SLIDE_RE.match(name)]
    return sorted(files, key=lambda name: int(SLIDE_RE.match(name).group(1)))


def relationship_targets(archive: zipfile.ZipFile, rels_path: str, base_dir: str) -> dict[str, str]:
    if rels_path not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read(rels_path))
    result: dict[str, str] = {}
    for rel in root.findall(f".//{{{PACKAGE_REL_NS}}}Relationship"):
        rel_id = rel.attrib.get("Id", "")
        target = rel.attrib.get("Target", "")
        if rel_id and target:
            result[rel_id] = posixpath.normpath(posixpath.join(base_dir, target))
    return result


def slide_text(archive: zipfile.ZipFile, slide_file: str) -> str:
    root = ET.fromstring(archive.read(slide_file))
    text_items = [clean_text(node.text, 300) for node in root.findall(".//a:t", DRAWING_NS)]
    return clean_text(" ".join(item for item in text_items if item), 5000)


def chart_type(root: ET.Element) -> str:
    for node in root.iter():
        if node.tag.startswith(f"{{{C_NS_URI}}}") and node.tag.endswith("Chart"):
            return node.tag.split("}", 1)[-1]
    return "unknown"


def cache_values(container: ET.Element | None) -> list[str]:
    if container is None:
        return []
    values: list[str] = []
    for node in container.findall(".//c:pt/c:v", CHART_NS):
        values.append(clean_text(node.text, 300))
    return values


def inspect_chart_xml(chart_bytes: bytes) -> dict[str, Any]:
    root = ET.fromstring(chart_bytes)
    series = root.findall(".//c:ser", CHART_NS)
    series_names: list[str] = []
    category_values: list[str] = []
    value_points: list[int] = []
    for ser in series:
        names = cache_values(ser.find("c:tx", CHART_NS))
        series_names.append(names[0] if names else "")
        if not category_values:
            category_values = cache_values(ser.find("c:cat", CHART_NS))
        value_points.append(len(cache_values(ser.find("c:val", CHART_NS))))
    return {
        "chart_type": chart_type(root),
        "series_count": len(series),
        "series_names": series_names,
        "category_values": category_values,
        "value_points": value_points,
    }


def embedded_workbooks(archive: zipfile.ZipFile, chart_path: str) -> list[str]:
    chart_name = posixpath.basename(chart_path)
    rels_path = posixpath.join(posixpath.dirname(chart_path), "_rels", chart_name + ".rels")
    rels = relationship_targets(archive, rels_path, posixpath.dirname(chart_path))
    return [
        path
        for path in rels.values()
        if path.startswith("ppt/embeddings/") and path.lower().endswith((".xlsx", ".xlsm", ".xlsb"))
    ]


def inspect_embedded_workbook(archive: zipfile.ZipFile, workbook_path: str) -> dict[str, Any]:
    if workbook_path not in archive.namelist():
        return {"status": "missing"}
    if workbook_path.lower().endswith(".xlsb"):
        return {"status": "unsupported_xlsb", "path": workbook_path}
    try:
        workbook = load_workbook(io.BytesIO(archive.read(workbook_path)), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        matrix: list[list[Any]] = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True):
            matrix.append([clean_text(value, 300) for value in row[:20]])
        first_row = [value for value in (matrix[0] if matrix else []) if value]
        first_col = [row[0] for row in matrix[1:] if row and row[0]]
        token_text = " ".join(str(value) for row in matrix for value in row if value)
        return {
            "status": "ok",
            "path": workbook_path,
            "sheet_name": sheet.title,
            "rows": sheet.max_row,
            "columns": sheet.max_column,
            "first_row": first_row[:30],
            "first_col": first_col[:30],
            "sample_matrix": matrix[:8],
            "tokens": sorted(tokenize(token_text))[:200],
            "shape_guess": guess_workbook_shape(first_row, first_col),
        }
    except Exception as exc:  # pragma: no cover - defensive for odd embedded files.
        return {"status": f"read_failed:{exc}", "path": workbook_path}


def guess_workbook_shape(first_row: list[str], first_col: list[str]) -> str:
    row_text = " ".join(first_row)
    col_text = " ".join(first_col)
    brand_keywords = ["百泽安", "达伯舒", "艾瑞卡", "拓益", "可瑞达", "欧狄沃", "汉斯状", "择捷美"]
    metric_keywords = ["总知晓", "处方", "nps", "覆盖", "频率", "第一提及", "tom", "sov", "soc"]
    row_brand = sum(1 for keyword in brand_keywords if keyword.lower() in row_text.lower())
    col_brand = sum(1 for keyword in brand_keywords if keyword.lower() in col_text.lower())
    row_metric = sum(1 for keyword in metric_keywords if keyword.lower() in row_text.lower())
    col_metric = sum(1 for keyword in metric_keywords if keyword.lower() in col_text.lower())
    if col_brand >= 2 and row_metric >= 1:
        return "metric_rows_brand_columns"
    if row_brand >= 2 and col_metric >= 1:
        return "brand_rows_metric_columns"
    if col_brand >= 2:
        return "brand_rows_single_metric"
    if row_brand >= 2:
        return "metric_rows_brand_columns"
    return "unknown"


def inspect_ppt_charts(pptx_path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with zipfile.ZipFile(pptx_path) as archive:
        for slide_file in ppt_slide_files(archive):
            slide_no = int(SLIDE_RE.match(slide_file).group(1))
            slide_root = ET.fromstring(archive.read(slide_file))
            rels = relationship_targets(
                archive,
                f"ppt/slides/_rels/slide{slide_no}.xml.rels",
                "ppt/slides",
            )
            text = slide_text(archive, slide_file)
            title = clean_text(text.split(" 来源：", 1)[0], 300)
            chart_sort = 0
            for chart in slide_root.findall(".//c:chart", {**CHART_NS, **REL_NS}):
                rel_id = chart.attrib.get(qn(R_NS_URI, "id"), "")
                if rel_id not in rels:
                    continue
                chart_sort += 1
                # ElementTree does not expose parent pointers, so locate the owning frame
                # explicitly; coordinates are useful when the workbook has generic labels.
                chart_x = chart_y = chart_cx = chart_cy = ""
                for candidate_frame in slide_root.findall(".//p:graphicFrame", {"p": "http://schemas.openxmlformats.org/presentationml/2006/main"}):
                    candidate_chart = candidate_frame.find(".//c:chart", {**CHART_NS, **REL_NS})
                    if candidate_chart is chart:
                        xfrm = candidate_frame.find(
                            ".//p:xfrm",
                            {"p": "http://schemas.openxmlformats.org/presentationml/2006/main", **DRAWING_NS},
                        )
                        if xfrm is not None:
                            off = xfrm.find("a:off", DRAWING_NS)
                            ext = xfrm.find("a:ext", DRAWING_NS)
                            if off is not None and ext is not None:
                                chart_x = off.attrib.get("x", "")
                                chart_y = off.attrib.get("y", "")
                                chart_cx = ext.attrib.get("cx", "")
                                chart_cy = ext.attrib.get("cy", "")
                        break
                chart_path = rels[rel_id]
                chart_info = inspect_chart_xml(archive.read(chart_path)) if chart_path in archive.namelist() else {}
                workbook_paths = embedded_workbooks(archive, chart_path)
                workbook_info = inspect_embedded_workbook(archive, workbook_paths[0]) if workbook_paths else {"status": "not_found"}
                combined_tokens = sorted(
                    tokenize(text)
                    | set(workbook_info.get("tokens", []))
                    | tokenize(" ".join(chart_info.get("series_names", [])))
                    | tokenize(" ".join(chart_info.get("category_values", [])))
                )
                rows.append(
                    {
                        "slide": slide_no,
                        "chart_sort": chart_sort,
                        "chart_path": chart_path,
                        "chart_x": chart_x,
                        "chart_y": chart_y,
                        "chart_cx": chart_cx,
                        "chart_cy": chart_cy,
                        "embedded_workbook": workbook_paths[0] if workbook_paths else "",
                        "embedded_status": workbook_info.get("status", ""),
                        "embedded_sheet": workbook_info.get("sheet_name", ""),
                        "embedded_rows": workbook_info.get("rows", ""),
                        "embedded_columns": workbook_info.get("columns", ""),
                        "chart_type": chart_info.get("chart_type", ""),
                        "series_count": chart_info.get("series_count", 0),
                        "series_names": chart_info.get("series_names", []),
                        "category_values": chart_info.get("category_values", []),
                        "workbook_first_row": workbook_info.get("first_row", []),
                        "workbook_first_col": workbook_info.get("first_col", []),
                        "workbook_sample_matrix": workbook_info.get("sample_matrix", []),
                        "shape_guess": workbook_info.get("shape_guess", "unknown"),
                        "slide_title": title,
                        "slide_text": text,
                        "tokens": combined_tokens[:300],
                    }
                )
    return rows


def write_chart_inventory(rows: list[dict[str, Any]], out_dir: Path) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "ppt_chart_inventory.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    csv_rows = []
    for row in rows:
        csv_rows.append(
            {
                "slide": row["slide"],
                "chart_sort": row["chart_sort"],
                "chart_path": row["chart_path"],
                "chart_x": row.get("chart_x", ""),
                "chart_y": row.get("chart_y", ""),
                "chart_cx": row.get("chart_cx", ""),
                "chart_cy": row.get("chart_cy", ""),
                "embedded_workbook": row["embedded_workbook"],
                "embedded_status": row["embedded_status"],
                "chart_type": row["chart_type"],
                "series_count": row["series_count"],
                "shape_guess": row["shape_guess"],
                "series_names": " | ".join(row.get("series_names", [])),
                "category_values": " | ".join(row.get("category_values", [])[:30]),
                "workbook_first_row": " | ".join(row.get("workbook_first_row", [])[:30]),
                "workbook_first_col": " | ".join(row.get("workbook_first_col", [])[:30]),
                "slide_title": row["slide_title"],
            }
        )
    write_csv(
        out_dir / "ppt_chart_inventory.csv",
        csv_rows,
        [
            "slide",
            "chart_sort",
            "chart_path",
            "chart_x",
            "chart_y",
            "chart_cx",
            "chart_cy",
            "embedded_workbook",
            "embedded_status",
            "chart_type",
            "series_count",
            "shape_guess",
            "series_names",
            "category_values",
            "workbook_first_row",
            "workbook_first_col",
            "slide_title",
        ],
    )
    status_counts: dict[str, int] = {}
    for row in rows:
        status = str(row.get("embedded_status", ""))
        status_counts[status] = status_counts.get(status, 0) + 1
    return {"charts": len(rows), "embedded_status_counts": status_counts}
