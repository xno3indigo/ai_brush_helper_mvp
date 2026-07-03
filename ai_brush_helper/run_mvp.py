#!/usr/bin/env python3
"""
AI brush helper MVP.

Two conservative workflows are supported:

1. mapping: map an existing PPT template to already-imported DP/database
   result tables using exported config tables.
2. infer-config: infer draft config tables when config tables do not exist yet.
3. extract: legacy context extraction for questionnaire/code-based brush-code
   generation.

The script does not write to the business database and does not modify PPTX
files. It produces reviewable mapping and validation reports first.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required. Run with a Python environment that has openpyxl installed."
    ) from exc


DRAWING_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
CHART_NS = {"c": "http://schemas.openxmlformats.org/drawingml/2006/chart"}
PLACEHOLDER_RE = re.compile(r"\$\{[^}]+\}")
SLIDE_RE = re.compile(r"ppt/slides/slide(\d+)\.xml$")
QUESTION_CODE_RE = re.compile(r"^([A-Za-z]+[A-Za-z0-9]*?(?:x?\d+)?(?:_[A-Za-z0-9]+)*)")
CONFIG_TABLE_NAMES = {"bh_database_table", "bh_database_table_field", "bh_charts_replaces"}
WAVE_SUFFIX_RE = re.compile(r"_\d{2}q[1-4]$", re.IGNORECASE)

MODULE_KEYWORDS = [
    ("brand_awareness", ["品牌知晓", "品牌知名", "awareness", "tom"]),
    ("adoption_curve", ["接纳阶梯", "曾经使用", "过去一个月使用", "最常使用", "adoption", "mopb"]),
    ("share_of_voice", ["品牌声量", "代表拜访", "被拜访", "拜访频次", "覆盖率", "share of voice", "call coverage", "call frequency", "sov"]),
    ("message_recall", ["关键信息", "自发回忆", "message recall", "key message", "k1", "k2", "k3"]),
    ("digital_platform", ["digital media", "数字平台", "线上活动", "platform"]),
    ("trial_awareness", ["临床试验", "trial", "familiarity"]),
    ("activity_performance", ["市场活动", "activity performance", "会议", "活动表现"]),
    ("brand_equity", ["品牌表现", "关键择药", "attributes", "整体满意度", "brand performance"]),
    ("brand_image", ["品牌形象", "推荐", "company", "nps"]),
]

PYTHON_BRUSH_PATTERNS = [
    "to_sql(",
    "bh_charts_replaces",
    "bh_database_table",
    "bh_database_table_field",
    "refresh_",
    "DELETE FROM",
    "VIEW_",
]

TABLE_PREFIX_BY_TOPIC = {
    "brand_awareness": ["brand_awareness"],
    "adoption_curve": ["adoption_curve"],
    "brand_equity": ["brand_equity"],
    "share_of_voice": ["share_of_voice"],
    "message_recall": ["message_recall"],
    "digital_platform": ["digital_platform"],
    "trial_awareness": ["trial_awareness", "trial"],
    "activity_performance": ["activity_performance"],
    "brand_image": ["brand_image"],
    "summary": ["summary", "patient_profile", "brand_share"],
}


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def normalize_placeholder(value: str) -> str:
    text = clean_one_line(value, 500)
    if text.startswith("${") and text.endswith("}"):
        inner = text[2:-1]
    else:
        inner = text
    inner = re.sub(r"\s+", "", inner)
    return "${" + inner + "}"


def clean_one_line(value: Any, limit: int = 500) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def normalize_key_map(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        if key is None:
            continue
        normalized[str(key).strip().lower()] = value
    return normalized


def row_get(row: dict[str, Any], *keys: str, default: Any = "") -> Any:
    lower = normalize_key_map(row)
    for key in keys:
        if key.lower() in lower:
            return lower[key.lower()]
    return default


def read_csv_table(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_xlsx_table(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_one_line(value, 200) for value in next(rows)]
    result: list[dict[str, Any]] = []
    for values in rows:
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if any(value not in (None, "") for value in row.values()):
            result.append(row)
    return result


def read_table_file(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        return read_csv_table(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_xlsx_table(path)
    raise ValueError(f"Unsupported table export: {path}")


def table_file_stem(path: Path) -> str:
    return path.stem.strip()


def normalize_table_name(value: str) -> str:
    return clean_one_line(value, 300).lower()


def strip_wave_suffix(value: str) -> str:
    return WAVE_SUFFIX_RE.sub("", clean_one_line(value, 300))


def placeholder_inner(value: str) -> str:
    normalized = normalize_placeholder(value)
    return normalized[2:-1]


def find_export_file(export_dir: Path, table_name: str) -> Path | None:
    candidates = []
    for suffix in [".csv", ".xlsx", ".xlsm"]:
        candidates.extend(export_dir.rglob(table_name + suffix))
    return candidates[0] if candidates else None


def load_db_exports(export_dir: Path) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[str]], dict[str, Any]]:
    config_tables: dict[str, list[dict[str, Any]]] = {}
    result_headers: dict[str, list[str]] = {}
    export_files: list[str] = []

    for table_name in CONFIG_TABLE_NAMES:
        path = find_export_file(export_dir, table_name)
        if path:
            config_tables[table_name] = read_table_file(path)
            export_files.append(str(path))
        else:
            config_tables[table_name] = []

    for path in sorted(export_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx", ".xlsm"}:
            continue
        stem = table_file_stem(path)
        if stem in CONFIG_TABLE_NAMES:
            continue
        try:
            rows = read_table_file(path)
        except Exception:
            continue
        if rows:
            result_headers[stem] = list(rows[0].keys())
        else:
            result_headers[stem] = []

    meta = {
        "db_export_dir": str(export_dir),
        "config_rows": {name: len(rows) for name, rows in config_tables.items()},
        "result_table_exports": len(result_headers),
        "config_files": export_files,
    }
    return config_tables, result_headers, meta


def resolve_result_table_name(table_name: str, result_headers: dict[str, list[str]]) -> str | None:
    if table_name in result_headers:
        return table_name
    target = normalize_table_name(table_name)
    for candidate in result_headers:
        if normalize_table_name(candidate) == target:
            return candidate
    return None


def ppt_slide_files(archive: zipfile.ZipFile) -> list[str]:
    files = [name for name in archive.namelist() if SLIDE_RE.match(name)]
    return sorted(files, key=lambda name: int(SLIDE_RE.match(name).group(1)))


def infer_module(text: str) -> str:
    low = text.lower()
    scored: list[tuple[int, str]] = []
    for module, keywords in MODULE_KEYWORDS:
        score = 0
        for keyword in keywords:
            keyword_low = keyword.lower()
            count = low.count(keyword_low)
            if count <= 0:
                continue
            score += count
            first_pos = low.find(keyword_low)
            if 0 <= first_pos < 350:
                score += 3
        if score:
            scored.append((score, module))
    if scored:
        scored.sort(key=lambda item: (-item[0], item[1]))
        return scored[0][1]
    return "unknown"


def extract_ppt_placeholders(pptx_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    placeholder_rows: list[dict[str, Any]] = []
    slide_rows: list[dict[str, Any]] = []

    with zipfile.ZipFile(pptx_path) as archive:
        total_chart_count = len([n for n in archive.namelist() if re.match(r"ppt/charts/chart\d+\.xml$", n)])
        embedding_count = len([n for n in archive.namelist() if n.startswith("ppt/embeddings/")])
        media_count = len([n for n in archive.namelist() if n.startswith("ppt/media/")])

        for slide_file in ppt_slide_files(archive):
            slide_no = int(SLIDE_RE.match(slide_file).group(1))
            root = ET.fromstring(archive.read(slide_file))
            text_items = [node.text or "" for node in root.findall(".//a:t", DRAWING_NS)]
            chart_count = len(root.findall(".//c:chart", CHART_NS))
            text_items = [clean_one_line(item, 300) for item in text_items if clean_one_line(item)]
            full_joined_text = clean_one_line(" ".join(text_items), 200000)
            report_joined_text = clean_one_line(full_joined_text, 1200)
            placeholders = PLACEHOLDER_RE.findall(full_joined_text)
            module = infer_module(full_joined_text)

            title = ""
            for item in text_items:
                if not PLACEHOLDER_RE.fullmatch(item):
                    title = item
                    break

            slide_rows.append(
                {
                    "slide": slide_no,
                    "module_guess": module,
                    "placeholder_count": len(placeholders),
                    "chart_count": chart_count,
                    "text_item_count": len(text_items),
                    "title_guess": title,
                    "sample_text": report_joined_text,
                }
            )

            for raw in placeholders:
                placeholder_rows.append(
                    {
                        "slide": slide_no,
                        "module_guess": module,
                        "placeholder_raw": raw,
                        "placeholder_norm": normalize_placeholder(raw),
                        "title_guess": title,
                        "nearby_text": report_joined_text,
                    }
                )

    meta = {
        "pptx": str(pptx_path),
        "slides": len(slide_rows),
        "charts": total_chart_count,
        "media": media_count,
        "embeddings": embedding_count,
        "placeholder_count": len(placeholder_rows),
        "unique_placeholder_count": len({row["placeholder_norm"] for row in placeholder_rows}),
    }
    return placeholder_rows, slide_rows, meta


def split_question_header(header: str) -> tuple[str, str]:
    header = clean_one_line(header, 2000)
    if " : " in header:
        left, right = header.split(" : ", 1)
        return left.strip(), right.strip()
    if ":" in header:
        left, right = header.split(":", 1)
        return left.strip(), right.strip()
    match = QUESTION_CODE_RE.match(header)
    return (match.group(1).strip(), header) if match else ("", header)


def question_prefix(question_code: str) -> str:
    match = re.match(r"^([A-Za-z]+)", question_code)
    return match.group(1).upper() if match else ""


def extract_excel_schema(xlsx_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_meta: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sheet_meta.append({"sheet": sheet_name, "rows": ws.max_row, "columns": ws.max_column})
        header_values = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        samples_by_col: dict[int, list[str]] = defaultdict(list)
        non_empty_by_col: Counter[int] = Counter()
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 12)):
            for index, cell in enumerate(row, start=1):
                if cell.value is not None and str(cell.value).strip() != "":
                    non_empty_by_col[index] += 1
                    if len(samples_by_col[index]) < 3:
                        samples_by_col[index].append(clean_one_line(cell.value, 80))

        for index, header in enumerate(header_values, start=1):
            raw_header = clean_one_line(header, 2000)
            if not raw_header:
                continue
            code, question_text = split_question_header(raw_header)
            rows.append(
                {
                    "sheet": sheet_name,
                    "column_index": index,
                    "column_letter": ws.cell(row=1, column=index).coordinate.replace("1", ""),
                    "raw_header": raw_header,
                    "question_code": code,
                    "question_prefix": question_prefix(code),
                    "question_text": question_text,
                    "sample_values": " | ".join(samples_by_col.get(index, [])),
                    "sample_non_empty_count_first_10": non_empty_by_col.get(index, 0),
                }
            )

    meta = {
        "xlsx": str(xlsx_path),
        "sheets": sheet_meta,
        "schema_rows": len(rows),
        "question_prefix_counts": dict(Counter(row["question_prefix"] for row in rows if row["question_prefix"])),
    }
    return rows, meta


def scan_existing_code(project_dir: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    module_counter: Counter[str] = Counter()
    file_counter: Counter[str] = Counter()

    for path in sorted(project_dir.rglob("*.py")):
        if "__pycache__" in path.parts:
            continue
        rel = path.relative_to(project_dir)
        try:
            lines = path.read_text(encoding="utf-8").splitlines()
        except UnicodeDecodeError:
            lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()

        current_class = ""
        current_func = ""
        for lineno, line in enumerate(lines, start=1):
            stripped = line.strip()
            class_match = re.match(r"class\s+([A-Za-z_][A-Za-z0-9_]*)", stripped)
            func_match = re.match(r"def\s+([A-Za-z_][A-Za-z0-9_]*)", stripped)
            if class_match:
                current_class = class_match.group(1)
            if func_match:
                current_func = func_match.group(1)

            matched = [pattern for pattern in PYTHON_BRUSH_PATTERNS if pattern in line]
            if matched:
                category = "refresh_or_template"
                if "to_sql(" in line:
                    category = "database_write"
                if "bh_database_table" in line or "bh_database_table_field" in line:
                    category = "chart_config"
                if "bh_charts_replaces" in line or "VIEW_" in line:
                    category = "ppt_variable"

                topic_guess = infer_topic_from_path(rel)
                rows.append(
                    {
                        "file": str(rel),
                        "line": lineno,
                        "class": current_class,
                        "function": current_func,
                        "category": category,
                        "topic_guess": topic_guess,
                        "matched": ",".join(matched),
                        "code": stripped[:500],
                    }
                )
                module_counter[topic_guess] += 1
                file_counter[str(rel)] += 1

    meta = {
        "project_dir": str(project_dir),
        "brush_points": len(rows),
        "topic_counts": dict(module_counter),
        "top_files": dict(file_counter.most_common(20)),
    }
    return rows, meta


def extract_table_patterns_from_line(line: str) -> list[str]:
    patterns: list[str] = []
    # Examples:
    # pd.read_sql_table(f"{topic}_total_{wave}", engine)
    # df.to_sql(name=f"{tbl_name}_{wave}", con=engine)
    # df.to_sql(f"{tbl_name}_{wave}", con=engine)
    regexes = [
        r"read_sql_table\(\s*f?[\"']([^\"']+)[\"']",
        r"\.to_sql\(\s*f?[\"']([^\"']+)[\"']",
        r"\.to_sql\(\s*name\s*=\s*f?[\"']([^\"']+)[\"']",
        r"SHOW TABLE STATUS LIKE\s+'([^']+)'",
    ]
    for regex in regexes:
        for match in re.finditer(regex, line):
            value = match.group(1).strip()
            if value and value not in {"bh_charts_replaces", "bh_database_table", "bh_database_table_field"}:
                patterns.append(value)
    return patterns


def scan_mapping_evidence(project_dir: Path) -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    evidence_rows: list[dict[str, Any]] = []
    table_patterns_by_func: dict[str, list[str]] = defaultdict(list)
    source_lines: list[dict[str, Any]] = []

    for path in sorted(project_dir.rglob("*.py")):
        if "__pycache__" in path.parts:
            continue
        rel = path.relative_to(project_dir)
        topic_guess = infer_topic_from_path(rel)
        try:
            lines = path.read_text(encoding="utf-8").splitlines()
        except UnicodeDecodeError:
            lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()

        current_func = ""
        global_vars: dict[str, str] = {}
        local_vars_by_func: dict[str, dict[str, str]] = defaultdict(dict)
        for lineno, line in enumerate(lines, start=1):
            stripped = line.strip()
            func_match = re.match(r"def\s+([A-Za-z_][A-Za-z0-9_]*)", stripped)
            if func_match:
                current_func = func_match.group(1)

            func_key = f"{rel}:{current_func}" if current_func else str(rel)
            assignment = re.match(r"(?P<var>[A-Za-z_][A-Za-z0-9_]*)\s*=\s*f?[\"'](?P<value>[^\"']+)[\"']", stripped)
            if assignment:
                target = local_vars_by_func[func_key] if current_func else global_vars
                target[assignment.group("var")] = substitute_simple_vars(assignment.group("value"), global_vars)

            var_scope = {**global_vars, **local_vars_by_func.get(func_key, {})}
            for table_pattern in extract_table_patterns_from_line(line):
                table_pattern = substitute_simple_vars(table_pattern, var_scope)
                if table_pattern not in table_patterns_by_func[func_key]:
                    table_patterns_by_func[func_key].append(table_pattern)

            source_lines.append(
                {
                    "file": str(rel),
                    "line": lineno,
                    "topic_guess": topic_guess,
                    "function": current_func,
                    "function_key": func_key,
                    "code": stripped,
                    "is_commented": stripped.startswith("#"),
                }
            )

    for item in source_lines:
        code = item["code"]
        code_without_comment = code[1:].strip() if item["is_commented"] else code

        # Direct refresh call: refresh_xxx(engine, 7, wave)
        for match in re.finditer(r"(?P<callee>refresh_[A-Za-z0-9_]+)\([^#\n]*?,\s*(?P<page>\d+)\s*,\s*wave", code_without_comment):
            callee = match.group("callee")
            page = int(match.group("page"))
            callee_key = find_function_key(table_patterns_by_func, item["file"], callee)
            evidence_rows.append(
                {
                    "page": page,
                    "code_topic": item["topic_guess"],
                    "relationship_type": "direct_refresh_call",
                    "active": not item["is_commented"],
                    "source_file": item["file"],
                    "source_line": item["line"],
                    "source_function": item["function"],
                    "callee": callee,
                    "candidate_table_patterns": " | ".join(table_patterns_by_func.get(callee_key, [])),
                    "evidence": code_without_comment[:500],
                }
            )

        # Tuple task pattern: (['east', ...], 177)
        tuple_match = re.search(r"\((?P<values>\[.*?\]|[A-Za-z_][A-Za-z0-9_]*),\s*(?P<page>\d{1,3})\)", code_without_comment)
        if tuple_match and "refresh_" in item["function"]:
            page = int(tuple_match.group("page"))
            evidence_rows.append(
                {
                    "page": page,
                    "code_topic": item["topic_guess"],
                    "relationship_type": "task_tuple",
                    "active": not item["is_commented"],
                    "source_file": item["file"],
                    "source_line": item["line"],
                    "source_function": item["function"],
                    "callee": "",
                    "candidate_table_patterns": "",
                    "evidence": code_without_comment[:500],
                }
            )

    return evidence_rows, table_patterns_by_func


def substitute_simple_vars(pattern: str, variables: dict[str, str]) -> str:
    out = pattern
    for key, value in variables.items():
        out = out.replace("{" + key + "}", value)
    return out


def find_function_key(table_patterns_by_func: dict[str, list[str]], source_file: str, function_name: str) -> str:
    suffix = f":{function_name}"
    same_file = f"{source_file}:{function_name}"
    if same_file in table_patterns_by_func:
        return same_file
    for key in table_patterns_by_func:
        if key.endswith(suffix):
            return key
    return same_file


def build_mapping_candidates(
    slide_rows: list[dict[str, Any]],
    evidence_rows: list[dict[str, Any]],
    wave: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    slide_by_page = {int(row["slide"]): row for row in slide_rows}
    candidates: list[dict[str, Any]] = []

    for evidence in evidence_rows:
        page = int(evidence["page"])
        slide = slide_by_page.get(page, {})
        slide_module = slide.get("module_guess", "unknown")
        code_topic = evidence["code_topic"]
        score = 0
        if evidence["active"]:
            score += 4
        if evidence["relationship_type"] == "direct_refresh_call":
            score += 3
        if slide_module == code_topic:
            score += 2
        if int(slide.get("placeholder_count", 0) or 0) > 0:
            score += 1
        if evidence["candidate_table_patterns"]:
            score += 1

        confidence = "low"
        if score >= 8:
            confidence = "high"
        elif score >= 5:
            confidence = "medium"

        candidates.append(
            {
                "page": page,
                "confidence": confidence,
                "score": score,
                "active_in_code": evidence["active"],
                "slide_module_guess": slide_module,
                "code_topic": code_topic,
                "relationship_type": evidence["relationship_type"],
                "candidate_table_patterns": materialize_wave_patterns(
                    evidence["candidate_table_patterns"],
                    wave,
                    evidence["code_topic"],
                ),
                "placeholder_count": slide.get("placeholder_count", ""),
                "title_guess": slide.get("title_guess", ""),
                "source_file": evidence["source_file"],
                "source_line": evidence["source_line"],
                "source_function": evidence["source_function"],
                "callee": evidence["callee"],
                "evidence": evidence["evidence"],
            }
        )

    # If a PPT page has placeholders but no code evidence, add a gap row for review.
    candidate_pages = {int(row["page"]) for row in candidates}
    for slide in slide_rows:
        page = int(slide["slide"])
        if int(slide["placeholder_count"]) <= 0 or page in candidate_pages:
            continue
        candidates.append(
            {
                "page": page,
                "confidence": "missing_code_evidence",
                "score": 0,
                "active_in_code": "",
                "slide_module_guess": slide["module_guess"],
                "code_topic": "",
                "relationship_type": "ppt_placeholder_without_code_mapping",
                "candidate_table_patterns": " | ".join(TABLE_PREFIX_BY_TOPIC.get(slide["module_guess"], [])),
                "placeholder_count": slide["placeholder_count"],
                "title_guess": slide["title_guess"],
                "source_file": "",
                "source_line": "",
                "source_function": "",
                "callee": "",
                "evidence": slide["sample_text"][:500],
            }
        )

    candidates.sort(key=lambda row: (int(row["page"]), -int(row["score"] or 0), str(row["source_file"])))
    meta = {
        "mapping_candidates": len(candidates),
        "high_confidence": sum(1 for row in candidates if row["confidence"] == "high"),
        "medium_confidence": sum(1 for row in candidates if row["confidence"] == "medium"),
        "missing_code_evidence": sum(1 for row in candidates if row["confidence"] == "missing_code_evidence"),
    }
    return candidates, meta


def materialize_wave_patterns(value: str, wave: str, topic: str) -> str:
    if not value:
        return value
    parts = []
    for item in value.split(" | "):
        out = item
        out = out.replace("{wave}", wave)
        out = out.replace("{last_wave}", "last_wave")
        out = out.replace("{topic}", topic)
        out = out.replace("{tbl_name}", "<tbl_name>")
        parts.append(out)
    return " | ".join(parts)


def safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(str(value).strip()))
    except ValueError:
        return default


def build_db_mapping_reports(
    slide_rows: list[dict[str, Any]],
    placeholder_rows: list[dict[str, Any]],
    config_tables: dict[str, list[dict[str, Any]]],
    result_headers: dict[str, list[str]],
    wave: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    db_tables = config_tables.get("bh_database_table", [])
    db_fields = config_tables.get("bh_database_table_field", [])
    replaces = config_tables.get("bh_charts_replaces", [])

    slide_by_page = {safe_int(row["slide"]): row for row in slide_rows}
    fields_by_table_ref: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for field in db_fields:
        refs = [
            row_get(field, "table_id", "tableId", "database_table_id", "databaseTableId", default=""),
            row_get(field, "table_name", "database_table_name", "databaseTableName", default=""),
        ]
        for ref in refs:
            ref_text = clean_one_line(ref, 300)
            if ref_text:
                fields_by_table_ref[ref_text].append(field)

    replace_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in replaces:
        name = clean_one_line(row_get(row, "name", "NAME", default=""), 300)
        if name:
            replace_by_name[normalize_placeholder(name)].append(row)

    mapping_rows: list[dict[str, Any]] = []
    missing_tables: list[dict[str, Any]] = []
    missing_columns: list[dict[str, Any]] = []
    missing_placeholders: list[dict[str, Any]] = []

    has_result_exports = bool(result_headers)
    for table_row in db_tables:
        table_id = clean_one_line(row_get(table_row, "id", "ID", "table_id", "tableId", default=""), 300)
        table_name = clean_one_line(
            row_get(table_row, "name", "NAME", "table_name", "database_table_name", default=""),
            300,
        )
        page = safe_int(row_get(table_row, "page", "PAGE", default=0))
        sort = safe_int(row_get(table_row, "sort", "SORT", default=0))
        group_sign = clean_one_line(row_get(table_row, "group_sign", "GROUP_SIGN", default=""), 300)
        pivot = row_get(table_row, "pivot", "PIVOT", default="")
        slide = slide_by_page.get(page, {})

        resolved_table_name = resolve_result_table_name(table_name, result_headers)
        if has_result_exports and not resolved_table_name:
            missing_tables.append(
                {
                    "page": page,
                    "sort": sort,
                    "db_table": table_name,
                    "reason": "result_table_export_not_found",
                    "source": "bh_database_table",
                }
            )

        fields: list[dict[str, Any]] = []
        seen_field_rows: set[int] = set()
        for ref in [table_id, table_name, group_sign]:
            ref_text = clean_one_line(ref, 300)
            if not ref_text:
                continue
            for field in fields_by_table_ref.get(ref_text, []):
                field_marker = id(field)
                if field_marker in seen_field_rows:
                    continue
                seen_field_rows.add(field_marker)
                fields.append(field)
        if not fields:
            mapping_rows.append(
                {
                    "page": page,
                    "ppt_object_type": "chart",
                    "ppt_key": f"chart_{sort}",
                    "db_table": table_name,
                    "sort": sort,
                    "group_sign": group_sign,
                    "pivot": pivot,
                    "db_column": "",
                    "ppt_field_name": "",
                    "replace_value": "",
                    "confidence": "high" if table_name else "low",
                    "status": "no_field_config",
                    "slide_module_guess": slide.get("module_guess", ""),
                    "slide_title_guess": slide.get("title_guess", ""),
                }
            )
            continue

        for field in fields:
            db_column = clean_one_line(
                row_get(field, "database_column_name", "DATABASE_COLUMN_NAME", "databaseColumnName", "column_name", default=""),
                300,
            )
            ppt_field_name = clean_one_line(row_get(field, "name", "NAME", "field_name", default=""), 300)
            status = "ok"
            if (
                has_result_exports
                and resolved_table_name
                and db_column
                and db_column not in result_headers[resolved_table_name]
            ):
                status = "missing_column"
                missing_columns.append(
                    {
                        "page": page,
                        "sort": sort,
                        "db_table": table_name,
                        "db_column": db_column,
                        "ppt_field_name": ppt_field_name,
                        "available_columns": " | ".join(result_headers.get(resolved_table_name, [])),
                    }
                )

            mapping_rows.append(
                {
                    "page": page,
                    "ppt_object_type": "chart",
                    "ppt_key": f"chart_{sort}",
                    "db_table": table_name,
                    "sort": sort,
                    "group_sign": group_sign,
                    "pivot": pivot,
                    "db_column": db_column,
                    "ppt_field_name": ppt_field_name,
                    "replace_value": "",
                    "confidence": "high",
                    "status": status,
                    "slide_module_guess": slide.get("module_guess", ""),
                    "slide_title_guess": slide.get("title_guess", ""),
                }
            )

    ppt_placeholders = sorted({row["placeholder_norm"] for row in placeholder_rows})
    placeholder_slide: dict[str, int] = {}
    placeholder_title: dict[str, str] = {}
    for row in placeholder_rows:
        placeholder_slide.setdefault(row["placeholder_norm"], safe_int(row["slide"]))
        placeholder_title.setdefault(row["placeholder_norm"], row.get("title_guess", ""))

    for placeholder in ppt_placeholders:
        page = placeholder_slide.get(placeholder, 0)
        matches = replace_by_name.get(placeholder, [])
        if not matches:
            missing_placeholders.append(
                {
                    "page": page,
                    "placeholder": placeholder,
                    "reason": "not_found_in_bh_charts_replaces",
                    "title_guess": placeholder_title.get(placeholder, ""),
                }
            )
            mapping_rows.append(
                {
                    "page": page,
                    "ppt_object_type": "text",
                    "ppt_key": placeholder,
                    "db_table": "",
                    "sort": "",
                    "group_sign": "",
                    "pivot": "",
                    "db_column": "",
                    "ppt_field_name": "",
                    "replace_value": "",
                    "confidence": "missing",
                    "status": "missing_placeholder_value",
                    "slide_module_guess": slide_by_page.get(page, {}).get("module_guess", ""),
                    "slide_title_guess": placeholder_title.get(placeholder, ""),
                }
            )
            continue
        for match in matches:
            value = row_get(match, "value", "VALUE", default="")
            mapping_rows.append(
                {
                    "page": page,
                    "ppt_object_type": "text",
                    "ppt_key": placeholder,
                    "db_table": "",
                    "sort": "",
                    "group_sign": "",
                    "pivot": "",
                    "db_column": "",
                    "ppt_field_name": "",
                    "replace_value": clean_one_line(value, 500),
                    "confidence": "high",
                    "status": "ok",
                    "slide_module_guess": slide_by_page.get(page, {}).get("module_guess", ""),
                    "slide_title_guess": placeholder_title.get(placeholder, ""),
                }
            )

    mapping_rows.sort(key=lambda row: (safe_int(row["page"]), str(row["ppt_object_type"]), safe_int(row["sort"])))
    meta = {
        "db_mapping_rows": len(mapping_rows),
        "db_chart_config_rows": len(db_tables),
        "db_field_config_rows": len(db_fields),
        "db_replace_rows": len(replaces),
        "missing_tables": len(missing_tables),
        "missing_columns": len(missing_columns),
        "missing_placeholders": len(missing_placeholders),
        "result_table_validation": "enabled" if has_result_exports else "skipped_no_result_table_exports",
        "wave": wave,
    }
    return mapping_rows, missing_placeholders, missing_tables, missing_columns, meta


def generate_mapping_spec_from_db(mapping_rows: list[dict[str, Any]], wave: str, spec_rules_path: Path | None = None) -> str:
    chart_rows = [row for row in mapping_rows if row["ppt_object_type"] == "chart"]
    text_rows = [row for row in mapping_rows if row["ppt_object_type"] == "text" and row["status"] == "ok"]
    by_page: dict[int, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: {"chart": [], "text": []})
    for row in chart_rows:
        by_page[safe_int(row["page"])]["chart"].append(row)
    for row in text_rows:
        by_page[safe_int(row["page"])]["text"].append(row)

    lines = [
        "# 根据数据库配置表生成的 PPT-DB 映射 spec",
        "# 这份文件可人工审查后作为渲染输入。",
        f"wave: {wave}",
        "inputs:",
        f"  spec_rules: {json.dumps(str(spec_rules_path) if spec_rules_path else '', ensure_ascii=False)}",
        "pages:",
    ]
    for page in sorted(by_page):
        page_data = by_page[page]
        lines.append(f"  - page: {page}")
        lines.append("    text_bindings:")
        seen_text = set()
        for row in page_data["text"]:
            key = row["ppt_key"]
            if key in seen_text:
                continue
            seen_text.add(key)
            lines.append(f"      - placeholder: {json.dumps(key, ensure_ascii=False)}")
            lines.append(f"        value: {json.dumps(row['replace_value'], ensure_ascii=False)}")
        if not seen_text:
            lines.append("      []")
        lines.append("    chart_bindings:")
        charts: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
        for row in page_data["chart"]:
            charts[(str(row["sort"]), row["db_table"])].append(row)
        if not charts:
            lines.append("      []")
        for (sort, table), rows in sorted(charts.items(), key=lambda item: (safe_int(item[0][0]), item[0][1])):
            first = rows[0]
            lines.append(f"      - sort: {safe_int(sort)}")
            lines.append(f"        table: {json.dumps(table, ensure_ascii=False)}")
            lines.append(f"        group_sign: {json.dumps(first['group_sign'], ensure_ascii=False)}")
            lines.append(f"        pivot: {json.dumps(str(first['pivot']), ensure_ascii=False)}")
            lines.append("        fields:")
            for row in rows:
                if not row["db_column"] and not row["ppt_field_name"]:
                    continue
                lines.append(f"          - db_column: {json.dumps(row['db_column'], ensure_ascii=False)}")
                lines.append(f"            ppt_name: {json.dumps(row['ppt_field_name'], ensure_ascii=False)}")
    return "\n".join(lines) + "\n"


def table_module_guess(table_name: str) -> str:
    lowered = normalize_table_name(table_name)
    for module, prefixes in TABLE_PREFIX_BY_TOPIC.items():
        if any(prefix in lowered for prefix in prefixes):
            return module
    return infer_module(lowered.replace("_", " "))


def tokenize_for_match(value: str) -> set[str]:
    lowered = normalize_table_name(value)
    tokens = set(re.findall(r"[a-z0-9]+|[\u4e00-\u9fff]+", lowered))
    stop_words = {"total", "dp", "view", "table", "data", "q1", "q2", "q3", "q4"}
    return {token for token in tokens if token and token not in stop_words and len(token) > 1}


def parse_spec_hints(spec_path: Path | None) -> dict[str, Any]:
    if not spec_path:
        return {"page_tables": {}, "text": ""}
    try:
        text = spec_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = spec_path.read_text(encoding="utf-8", errors="ignore")

    page_tables: dict[int, set[str]] = defaultdict(set)
    current_page = 0
    for raw_line in text.splitlines():
        line = raw_line.strip()
        page_match = re.match(r"-?\s*page\s*:\s*(\d+)", line)
        if page_match:
            current_page = int(page_match.group(1))
            continue
        if current_page <= 0:
            continue
        for table_name in re.findall(r"[A-Za-z][A-Za-z0-9_]*_\d{2}q[1-4]", line):
            page_tables[current_page].add(table_name)

    return {"page_tables": {page: sorted(tables) for page, tables in page_tables.items()}, "text": text}


def score_table_for_slide(slide: dict[str, Any], table_name: str, wave: str, spec_hints: dict[str, Any]) -> tuple[int, list[str]]:
    score = 0
    reasons: list[str] = []
    page = safe_int(slide.get("slide"))
    slide_module = slide.get("module_guess", "unknown")
    table_module = table_module_guess(table_name)
    table_lower = normalize_table_name(table_name)

    page_hint_tables = set(spec_hints.get("page_tables", {}).get(page, []))
    if table_name in page_hint_tables:
        score += 12
        reasons.append("spec_page_expected_table")

    if wave and wave.lower() in table_lower:
        score += 2
        reasons.append("wave_suffix_match")

    if slide_module != "unknown" and table_module == slide_module:
        score += 7
        reasons.append("module_match")
    elif slide_module != "unknown":
        prefixes = TABLE_PREFIX_BY_TOPIC.get(slide_module, [])
        if any(prefix in table_lower for prefix in prefixes):
            score += 5
            reasons.append("module_prefix_match")

    slide_text = " ".join(
        [
            clean_one_line(slide.get("title_guess", ""), 500),
            clean_one_line(slide.get("sample_text", ""), 1200),
        ]
    )
    slide_tokens = tokenize_for_match(slide_text)
    table_tokens = tokenize_for_match(strip_wave_suffix(table_name).replace("_", " "))
    overlap = sorted(slide_tokens & table_tokens)
    if overlap:
        overlap_score = min(4, len(overlap))
        score += overlap_score
        reasons.append("token_overlap:" + ",".join(overlap[:6]))

    if safe_int(slide.get("chart_count")) > 0:
        score += 1
        reasons.append("slide_has_chart")

    if "total" in table_lower and re.search(r"\btotal\b|总体|总样本|全部", slide_text, re.IGNORECASE):
        score += 1
        reasons.append("total_context_match")

    return score, reasons


def confidence_from_score(score: int) -> str:
    if score >= 12:
        return "high"
    if score >= 8:
        return "medium"
    if score >= 5:
        return "low"
    return "no_match"


def field_display_name(column_name: str, history_field_name: str = "") -> str:
    if history_field_name:
        return history_field_name
    cleaned = clean_one_line(column_name, 200)
    replacements = {
        "brand": "品牌",
        "value": "数值",
        "name": "名称",
        "group": "分组",
        "segment": "分组",
        "base": "Base",
        "n": "N",
        "pct": "百分比",
        "percent": "百分比",
    }
    return replacements.get(cleaned.lower(), cleaned)


def build_history_indexes(config_tables: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    tables = config_tables.get("bh_database_table", [])
    fields = config_tables.get("bh_database_table_field", [])
    table_by_base: dict[str, dict[str, Any]] = {}
    fields_by_table_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    field_name_by_base_and_column: dict[tuple[str, str], str] = {}

    for row in tables:
        table_id = clean_one_line(row_get(row, "id", "ID", "table_id", default=""), 200)
        table_name = clean_one_line(row_get(row, "name", "NAME", "table_name", default=""), 300)
        group_sign = clean_one_line(row_get(row, "group_sign", "GROUP_SIGN", default=""), 300)
        base_key = strip_wave_suffix(table_name or group_sign)
        if base_key:
            table_by_base[base_key] = row
        if table_id:
            fields_by_table_id[table_id] = []

    for field in fields:
        table_id = clean_one_line(row_get(field, "table_id", "tableId", "database_table_id", default=""), 200)
        if table_id:
            fields_by_table_id[table_id].append(field)

    for row in tables:
        table_id = clean_one_line(row_get(row, "id", "ID", "table_id", default=""), 200)
        table_name = clean_one_line(row_get(row, "name", "NAME", "table_name", default=""), 300)
        base_key = strip_wave_suffix(table_name)
        for field in fields_by_table_id.get(table_id, []):
            column = clean_one_line(
                row_get(field, "database_column_name", "DATABASE_COLUMN_NAME", "column_name", default=""),
                200,
            )
            display = clean_one_line(row_get(field, "name", "NAME", "field_name", default=""), 200)
            if base_key and column and display:
                field_name_by_base_and_column[(base_key, column)] = display

    return {
        "table_by_base": table_by_base,
        "field_name_by_base_and_column": field_name_by_base_and_column,
    }


def infer_config_tables(
    slide_rows: list[dict[str, Any]],
    placeholder_rows: list[dict[str, Any]],
    result_headers: dict[str, list[str]],
    wave: str,
    spec_hints: dict[str, Any],
    history_config_tables: dict[str, list[dict[str, Any]]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    history_indexes = build_history_indexes(history_config_tables or {})
    history_tables = history_indexes["table_by_base"]
    history_field_names = history_indexes["field_name_by_base_and_column"]

    generated_tables: list[dict[str, Any]] = []
    generated_fields: list[dict[str, Any]] = []
    generated_replaces: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []

    table_id = 1
    used_page_table: set[tuple[int, str]] = set()
    result_table_names = sorted(result_headers)
    page_hint_tables: dict[int, list[str]] = spec_hints.get("page_tables", {})

    for slide in slide_rows:
        page = safe_int(slide.get("slide"))
        chart_count = safe_int(slide.get("chart_count"))
        if chart_count <= 0:
            continue

        scored: list[dict[str, Any]] = []
        for table_name in result_table_names:
            score, reasons = score_table_for_slide(slide, table_name, wave, spec_hints)
            confidence = confidence_from_score(score)
            if confidence == "no_match" and table_name not in page_hint_tables.get(page, []):
                continue
            scored.append(
                {
                    "table_name": table_name,
                    "score": score,
                    "confidence": confidence,
                    "reasons": reasons,
                }
            )

        scored.sort(key=lambda row: (-safe_int(row["score"]), row["table_name"]))
        selected = scored[: max(1, min(chart_count, len(scored)))]
        for index, item in enumerate(selected, start=1):
            table_name = item["table_name"]
            if (page, table_name) in used_page_table:
                continue
            used_page_table.add((page, table_name))
            base_key = strip_wave_suffix(table_name)
            history_row = history_tables.get(base_key, {})
            group_sign = clean_one_line(row_get(history_row, "group_sign", "GROUP_SIGN", default=""), 300) or base_key
            pivot = row_get(history_row, "pivot", "PIVOT", default="0")
            sort = safe_int(row_get(history_row, "sort", "SORT", default=index), index)
            status = "ready" if item["confidence"] == "high" else "needs_review"

            generated_tables.append(
                {
                    "id": table_id,
                    "page": page,
                    "sort": sort,
                    "name": table_name,
                    "group_sign": group_sign,
                    "pivot": pivot,
                    "confidence": item["confidence"],
                    "status": status,
                    "reason": ";".join(item["reasons"]),
                }
            )

            headers = result_headers.get(table_name, [])
            if not headers:
                review_status = "missing_result_columns"
            else:
                review_status = status
            review_rows.append(
                {
                    "page": page,
                    "slide_module_guess": slide.get("module_guess", ""),
                    "slide_title_guess": slide.get("title_guess", ""),
                    "chart_sort": sort,
                    "candidate_table": table_name,
                    "score": item["score"],
                    "confidence": item["confidence"],
                    "status": review_status,
                    "reason": ";".join(item["reasons"]),
                    "result_columns": " | ".join(headers),
                    "sample_text": slide.get("sample_text", ""),
                }
            )

            for column in headers:
                display = field_display_name(column, history_field_names.get((base_key, column), ""))
                generated_fields.append(
                    {
                        "table_id": table_id,
                        "name": display,
                        "database_column_name": column,
                        "confidence": "medium" if (base_key, column) in history_field_names else "low",
                        "status": "ready" if (base_key, column) in history_field_names else "needs_review",
                        "reason": "copied_from_history" if (base_key, column) in history_field_names else "from_result_table_header",
                    }
                )
            table_id += 1

    seen_placeholders: set[str] = set()
    for row in placeholder_rows:
        placeholder = row["placeholder_norm"]
        if placeholder in seen_placeholders:
            continue
        seen_placeholders.add(placeholder)
        generated_replaces.append(
            {
                "name": placeholder_inner(placeholder),
                "value": "TODO_REVIEW",
                "page": row.get("slide", ""),
                "confidence": "low",
                "status": "needs_value",
                "reason": "ppt_placeholder_found_value_requires_review",
            }
        )

    meta = {
        "generated_database_table_rows": len(generated_tables),
        "generated_database_table_field_rows": len(generated_fields),
        "generated_charts_replaces_rows": len(generated_replaces),
        "review_rows": len(review_rows),
        "high_confidence_tables": sum(1 for row in generated_tables if row["confidence"] == "high"),
        "medium_confidence_tables": sum(1 for row in generated_tables if row["confidence"] == "medium"),
        "low_confidence_tables": sum(1 for row in generated_tables if row["confidence"] == "low"),
        "result_table_exports": len(result_headers),
        "wave": wave,
    }
    return generated_tables, generated_fields, generated_replaces, review_rows, meta


def generate_inferred_config_spec(
    generated_tables: list[dict[str, Any]],
    generated_fields: list[dict[str, Any]],
    generated_replaces: list[dict[str, Any]],
    wave: str,
    spec_rules_path: Path | None = None,
) -> str:
    fields_by_table_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for field in generated_fields:
        fields_by_table_id[str(field["table_id"])].append(field)
    replaces_by_page: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in generated_replaces:
        replaces_by_page[safe_int(row.get("page"))].append(row)

    tables_by_page: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in generated_tables:
        tables_by_page[safe_int(row["page"])].append(row)

    lines = [
        "# 自动推断生成的配置表草稿 spec",
        "# 低置信度和 TODO_REVIEW 必须人工确认后才能导入数据库。",
        f"wave: {wave}",
        "inputs:",
        f"  spec_rules: {json.dumps(str(spec_rules_path) if spec_rules_path else '', ensure_ascii=False)}",
        "pages:",
    ]
    for page in sorted(set(tables_by_page) | set(replaces_by_page)):
        lines.append(f"  - page: {page}")
        lines.append("    text_replaces:")
        if not replaces_by_page.get(page):
            lines.append("      []")
        for row in replaces_by_page.get(page, []):
            lines.append(f"      - name: {json.dumps(row['name'], ensure_ascii=False)}")
            lines.append(f"        value: {json.dumps(row['value'], ensure_ascii=False)}")
            lines.append(f"        status: {row['status']}")
        lines.append("    chart_configs:")
        if not tables_by_page.get(page):
            lines.append("      []")
        for table in sorted(tables_by_page.get(page, []), key=lambda row: safe_int(row["sort"])):
            lines.append(f"      - id: {table['id']}")
            lines.append(f"        sort: {table['sort']}")
            lines.append(f"        table: {json.dumps(table['name'], ensure_ascii=False)}")
            lines.append(f"        group_sign: {json.dumps(table['group_sign'], ensure_ascii=False)}")
            lines.append(f"        pivot: {json.dumps(str(table['pivot']), ensure_ascii=False)}")
            lines.append(f"        confidence: {table['confidence']}")
            lines.append(f"        status: {table['status']}")
            lines.append("        fields:")
            for field in fields_by_table_id.get(str(table["id"]), []):
                lines.append(f"          - db_column: {json.dumps(field['database_column_name'], ensure_ascii=False)}")
                lines.append(f"            ppt_name: {json.dumps(field['name'], ensure_ascii=False)}")
                lines.append(f"            status: {field['status']}")
    return "\n".join(lines) + "\n"


def infer_topic_from_path(path: Path) -> str:
    name = path.stem.lower()
    mapping = {
        "brandawareness": "brand_awareness",
        "adoptioncurve": "adoption_curve",
        "brandequity": "brand_equity",
        "sov": "share_of_voice",
        "messagerecall": "message_recall",
        "trialawareness": "trial_awareness",
        "activityperformance": "activity_performance",
        "digitalplatform": "digital_platform",
        "brandimage": "brand_image",
        "refresh_share_of_voice": "share_of_voice",
        "refresh_message_recall_var": "message_recall",
        "refresh_adoption_curve_var": "adoption_curve",
        "refresh_brand_equity_var": "brand_equity",
        "refresh_trial_var": "trial_awareness",
        "refresh_summary_var": "summary",
    }
    return mapping.get(name, name)


def generate_initial_spec(
    slide_rows: list[dict[str, Any]],
    placeholder_rows: list[dict[str, Any]],
    excel_meta: dict[str, Any],
    code_meta: dict[str, Any],
    wave: str,
) -> str:
    placeholders_by_slide: dict[int, list[str]] = defaultdict(list)
    for row in placeholder_rows:
        placeholders_by_slide[int(row["slide"])].append(row["placeholder_norm"])

    lines: list[str] = []
    lines.append("# AI brush helper initial spec")
    lines.append("# Review this file before generating brush code.")
    lines.append(f"wave: {wave}")
    lines.append("inputs:")
    lines.append(f"  excel: {excel_meta['xlsx']}")
    lines.append(f"  project_dir: {code_meta['project_dir']}")
    lines.append("base_candidates:")
    lines.append("  total: null  # fill from questionnaire total sample or table spec")
    lines.append("  long: null   # fill when module should use long questionnaire")
    lines.append("pages:")

    for slide in slide_rows:
        slide_no = int(slide["slide"])
        if slide["placeholder_count"] == 0:
            continue
        placeholders = sorted(set(placeholders_by_slide.get(slide_no, [])))
        lines.append(f"  - page: {slide_no}")
        lines.append(f"    module_guess: {slide['module_guess']}")
        lines.append(f"    title_guess: {json.dumps(slide['title_guess'], ensure_ascii=False)}")
        lines.append("    base: TODO")
        lines.append("    status: needs_review")
        lines.append("    source_questions: []")
        lines.append("    output_tables: []")
        lines.append("    ppt_variables:")
        for item in placeholders[:80]:
            lines.append(f"      - name: {json.dumps(item, ensure_ascii=False)}")
            lines.append("        value: TODO")
        if len(placeholders) > 80:
            lines.append(f"      # ... {len(placeholders) - 80} more placeholders on this page")
        lines.append("    notes:")
        lines.append(f"      - {json.dumps(slide['sample_text'][:220], ensure_ascii=False)}")

    return "\n".join(lines) + "\n"


def build_summary(
    ppt_meta: dict[str, Any],
    excel_meta: dict[str, Any] | None,
    code_meta: dict[str, Any] | None,
    mapping_meta: dict[str, Any],
    placeholder_rows: list[dict[str, Any]],
    slide_rows: list[dict[str, Any]],
    db_export_meta: dict[str, Any] | None = None,
    mode: str = "extract",
) -> dict[str, Any]:
    module_pages: dict[str, list[int]] = defaultdict(list)
    for row in slide_rows:
        if int(row["placeholder_count"]) > 0:
            module_pages[row["module_guess"]].append(int(row["slide"]))
    return {
        "mode": mode,
        "ppt": ppt_meta,
        "excel": excel_meta or {},
        "code": code_meta or {},
        "db_exports": db_export_meta or {},
        "mapping": mapping_meta,
        "placeholder_module_counts": dict(Counter(row["module_guess"] for row in placeholder_rows)),
        "pages_with_placeholders_by_module": {k: v for k, v in sorted(module_pages.items())},
    }


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create PPT-to-DB mapping and brush-data extraction reports.")
    parser.add_argument(
        "--mode",
        choices=["mapping", "infer-config", "extract"],
        default="mapping",
        help=(
            "mapping: validate existing config tables. "
            "infer-config: generate draft config tables from PPT/result tables. "
            "extract: legacy PPT + Excel + old code reports."
        ),
    )
    parser.add_argument("--pptx", required=True, type=Path, help="Template PPTX path.")
    parser.add_argument("--excel", type=Path, help="Questionnaire Excel path; required only in extract mode.")
    parser.add_argument("--project", type=Path, help="Existing brush-code project path; required only in extract mode.")
    parser.add_argument("--db-export-dir", type=Path, help="Directory containing DB config/result table exports.")
    parser.add_argument("--history-config-dir", type=Path, help="Optional previous-wave config exports for infer-config mode.")
    parser.add_argument("--spec", type=Path, help="Optional reviewed spec/rules file for traceability.")
    parser.add_argument("--wave", default="25q3", help="Target wave label.")
    parser.add_argument("--out", required=True, type=Path, help="Output report directory.")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    ensure_dir(args.out)

    if not args.pptx.exists():
        raise SystemExit(f"Missing pptx: {args.pptx}")
    if args.spec and not args.spec.exists():
        raise SystemExit(f"Missing spec: {args.spec}")
    if args.history_config_dir and not args.history_config_dir.exists():
        raise SystemExit(f"Missing history_config_dir: {args.history_config_dir}")
    if args.mode in {"mapping", "infer-config"}:
        if not args.db_export_dir:
            raise SystemExit(f"Missing --db-export-dir in {args.mode} mode.")
        if not args.db_export_dir.exists():
            raise SystemExit(f"Missing db_export_dir: {args.db_export_dir}")
    if args.mode == "extract":
        for path_name in ["excel", "project"]:
            path = getattr(args, path_name)
            if path is None:
                raise SystemExit(f"Missing --{path_name} in extract mode.")
            if not path.exists():
                raise SystemExit(f"Missing {path_name}: {path}")

    placeholder_rows, slide_rows, ppt_meta = extract_ppt_placeholders(args.pptx)
    write_csv(
        args.out / "ppt_placeholders.csv",
        placeholder_rows,
        ["slide", "module_guess", "placeholder_raw", "placeholder_norm", "title_guess", "nearby_text"],
    )
    write_csv(
        args.out / "ppt_slides.csv",
        slide_rows,
        ["slide", "module_guess", "placeholder_count", "chart_count", "text_item_count", "title_guess", "sample_text"],
    )

    if args.mode == "mapping":
        config_tables, result_headers, db_export_meta = load_db_exports(args.db_export_dir)
        (
            db_mapping_rows,
            missing_placeholders,
            missing_tables,
            missing_columns,
            mapping_meta,
        ) = build_db_mapping_reports(slide_rows, placeholder_rows, config_tables, result_headers, args.wave)

        write_csv(
            args.out / "ppt_to_db_mapping.csv",
            db_mapping_rows,
            [
                "page",
                "ppt_object_type",
                "ppt_key",
                "db_table",
                "sort",
                "group_sign",
                "pivot",
                "db_column",
                "ppt_field_name",
                "replace_value",
                "confidence",
                "status",
                "slide_module_guess",
                "slide_title_guess",
            ],
        )
        write_csv(
            args.out / "missing_placeholders.csv",
            missing_placeholders,
            ["page", "placeholder", "reason", "title_guess"],
        )
        write_csv(
            args.out / "missing_tables.csv",
            missing_tables,
            ["page", "sort", "db_table", "reason", "source"],
        )
        write_csv(
            args.out / "missing_columns.csv",
            missing_columns,
            ["page", "sort", "db_table", "db_column", "ppt_field_name", "available_columns"],
        )

        spec = generate_mapping_spec_from_db(db_mapping_rows, args.wave, args.spec)
        (args.out / "mapping_spec.generated.yaml").write_text(spec, encoding="utf-8")

        summary = build_summary(
            ppt_meta,
            None,
            None,
            mapping_meta,
            placeholder_rows,
            slide_rows,
            db_export_meta,
            mode=args.mode,
        )
        (args.out / "summary.json").write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    if args.mode == "infer-config":
        _config_tables, result_headers, db_export_meta = load_db_exports(args.db_export_dir)
        history_config_tables: dict[str, list[dict[str, Any]]] = {}
        history_meta: dict[str, Any] = {}
        if args.history_config_dir:
            history_config_tables, _history_result_headers, history_meta = load_db_exports(args.history_config_dir)

        spec_hints = parse_spec_hints(args.spec)
        (
            generated_tables,
            generated_fields,
            generated_replaces,
            review_rows,
            infer_meta,
        ) = infer_config_tables(
            slide_rows,
            placeholder_rows,
            result_headers,
            args.wave,
            spec_hints,
            history_config_tables,
        )

        write_csv(
            args.out / "bh_database_table.generated.csv",
            generated_tables,
            ["id", "page", "sort", "name", "group_sign", "pivot", "confidence", "status", "reason"],
        )
        write_csv(
            args.out / "bh_database_table_field.generated.csv",
            generated_fields,
            ["table_id", "name", "database_column_name", "confidence", "status", "reason"],
        )
        write_csv(
            args.out / "bh_charts_replaces.generated.csv",
            generated_replaces,
            ["name", "value", "page", "confidence", "status", "reason"],
        )
        write_csv(
            args.out / "mapping_review.csv",
            review_rows,
            [
                "page",
                "slide_module_guess",
                "slide_title_guess",
                "chart_sort",
                "candidate_table",
                "score",
                "confidence",
                "status",
                "reason",
                "result_columns",
                "sample_text",
            ],
        )

        spec = generate_inferred_config_spec(generated_tables, generated_fields, generated_replaces, args.wave, args.spec)
        (args.out / "config_spec.generated.yaml").write_text(spec, encoding="utf-8")

        validation_report = {
            "mode": args.mode,
            "must_review": {
                "low_confidence_table_rows": infer_meta["low_confidence_tables"],
                "placeholder_values_todo": sum(1 for row in generated_replaces if row["value"] == "TODO_REVIEW"),
                "field_names_need_review": sum(1 for row in generated_fields if row["status"] == "needs_review"),
            },
            "notes": [
                "Generated config tables are drafts. Review before importing into database.",
                "bh_charts_replaces.generated.csv uses TODO_REVIEW where the value cannot be inferred from result-table headers.",
                "Use mapping mode after editing/importing generated config tables to validate them against the PPT and result tables.",
            ],
            "spec_page_tables": spec_hints.get("page_tables", {}),
            "history_config": history_meta,
        }
        (args.out / "validation_report.json").write_text(
            json.dumps(validation_report, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        infer_meta = {
            **infer_meta,
            "spec_page_hints": sum(len(tables) for tables in spec_hints.get("page_tables", {}).values()),
            "history_config_rows": history_meta.get("config_rows", {}),
        }
        summary = build_summary(
            ppt_meta,
            None,
            None,
            infer_meta,
            placeholder_rows,
            slide_rows,
            db_export_meta,
            mode=args.mode,
        )
        (args.out / "summary.json").write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0

    excel_rows, excel_meta = extract_excel_schema(args.excel)
    code_rows, code_meta = scan_existing_code(args.project)
    mapping_evidence_rows, _table_patterns_by_func = scan_mapping_evidence(args.project)
    mapping_candidate_rows, mapping_meta = build_mapping_candidates(slide_rows, mapping_evidence_rows, args.wave)

    write_csv(
        args.out / "excel_schema.csv",
        excel_rows,
        [
            "sheet",
            "column_index",
            "column_letter",
            "raw_header",
            "question_code",
            "question_prefix",
            "question_text",
            "sample_values",
            "sample_non_empty_count_first_10",
        ],
    )
    write_csv(
        args.out / "code_brush_points.csv",
        code_rows,
        ["file", "line", "class", "function", "category", "topic_guess", "matched", "code"],
    )
    write_csv(
        args.out / "mapping_candidates.csv",
        mapping_candidate_rows,
        [
            "page",
            "confidence",
            "score",
            "active_in_code",
            "slide_module_guess",
            "code_topic",
            "relationship_type",
            "candidate_table_patterns",
            "placeholder_count",
            "title_guess",
            "source_file",
            "source_line",
            "source_function",
            "callee",
            "evidence",
        ],
    )

    spec = generate_initial_spec(slide_rows, placeholder_rows, excel_meta, code_meta, args.wave)
    (args.out / "initial_spec.yaml").write_text(spec, encoding="utf-8")

    summary = build_summary(ppt_meta, excel_meta, code_meta, mapping_meta, placeholder_rows, slide_rows, mode=args.mode)
    (args.out / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
